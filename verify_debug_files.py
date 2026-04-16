#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
验证调试下载的 xlsx 是否有数据，同时修复 download_only.py 的 parse 逻辑：
解析前先校验文件头是否为 PK（真正 xlsx），否则不解析
"""
import os
from pathlib import Path
import openpyxl

DEBUG_DIR = Path(__file__).parent / "output" / "downloads_debug"

for f in sorted(DEBUG_DIR.glob("*.bin")):
    print(f"\n文件: {f.name}  ({f.stat().st_size} bytes)")
    with open(f, 'rb') as fh:
        header = fh.read(2)
    if header != b'PK':
        print("  ❌ 非 xlsx 格式，跳过")
        continue
    try:
        wb = openpyxl.load_workbook(str(f), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
        print(f"  ✅ 解析成功，前5行:")
        for r in rows:
            print(f"    {r}")
        total = ws.max_row - 1
        print(f"  总数据行: {total}")
        wb.close()
    except Exception as e:
        print(f"  解析异常: {e}")
