#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
YP 增量同步脚本 v1
=========================================================
用途：全量抓取完成后，定期轮询 YP 平台的变化：
  1. 新增商户  → 全量下载其商品
  2. 已有商户  → 对比商品列表，发现新增/删除的商品
  3. 差异结果  → 写入 MySQL（yp_products）+ 追加 Excel

与 download_only.py 的区别：
  - download_only.py：初次全量采集，商品全存 state.json（内存压力大）
  - 本脚本：增量模式，商品数据直接读写 MySQL，state.json 只存商户级快照
    state 结构：{ mid: { "count": N, "asin_set": [...], "last_sync": "ISO时间" } }

运行方式：
  python -X utf8 yp_incremental_sync.py              # 正常增量
  python -X utf8 yp_incremental_sync.py --force-all  # 强制重新对比所有商户
  python -X utf8 yp_incremental_sync.py --new-only   # 只处理本次新发现的商户
=========================================================
"""
import argparse
import json
import os
import re
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path

# ─── 配置 ──────────────────────────────────────────────────────────────────
SITE_ID      = "12002"
BASE_URL     = "https://www.yeahpromos.com"
SCRIPT_DIR   = Path(__file__).parent.resolve()
OUTPUT_DIR   = str(SCRIPT_DIR / "output")
DOWNLOAD_DIR = os.path.join(OUTPUT_DIR, "downloads_incr")
SYNC_STATE   = os.path.join(OUTPUT_DIR, "incremental_state.json")
LOG_FILE     = os.path.join(OUTPUT_DIR, "incremental_log.txt")
EXCEL_OUT    = r"D:\workspace\YP_incremental.xlsx"   # 增量变化单独一个文件

# MySQL 连接（与 download_only 写同一个库）
MYSQL_CFG = dict(host="localhost", port=3306, user="root",
                 password="admin", database="affiliate_marketing",
                 charset="utf8mb4")

COL_KEYS  = ["merchant_name","merchant_id","asin","product_name",
             "category","price","commission","tracking_link","scraped_at"]
COL_NAMES = ["商户名称","商户ID","ASIN","商品名称",
             "类别","价格(USD)","佣金率","投放链接","采集时间"]

# ─── 日志 ──────────────────────────────────────────────────────────────────
_log_fh = None

def _get_log_fh():
    global _log_fh
    if _log_fh is None:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        _log_fh = open(LOG_FILE, 'a', encoding='utf-8', buffering=1)
    return _log_fh

def log(tag, msg):
    ts  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}][{tag}] {msg}"
    print(line, flush=True)
    try:
        _get_log_fh().write(line + "\n")
    except Exception:
        pass

# ─── 增量状态管理 ──────────────────────────────────────────────────────────
def load_sync_state():
    """
    返回 dict: { mid(str): {"count": N, "asins": [list], "last_sync": "ISO"} }
    """
    if Path(SYNC_STATE).exists():
        with open(SYNC_STATE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_sync_state(state: dict):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    tmp = SYNC_STATE + ".tmp"
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(state, f, ensure_ascii=False, indent=2)
    os.replace(tmp, SYNC_STATE)

# ─── MySQL ─────────────────────────────────────────────────────────────────
def get_db_conn():
    import mysql.connector
    return mysql.connector.connect(**MYSQL_CFG)

def get_existing_asins_from_db(mid: str) -> set:
    """从 MySQL 查询某商户现有的 ASIN 集合"""
    try:
        conn = get_db_conn()
        cur  = conn.cursor()
        cur.execute("SELECT asin FROM yp_products WHERE merchant_id=%s", (mid,))
        rows = cur.fetchall()
        cur.close(); conn.close()
        return {r[0] for r in rows}
    except Exception as e:
        log("DB", f"查询商户 {mid} ASIN 失败: {e}")
        return set()

def insert_products_to_db(products: list):
    """插入新商品到 MySQL（ON DUPLICATE KEY UPDATE 幂等）"""
    if not products:
        return 0
    try:
        conn = get_db_conn()
        cur  = conn.cursor()
        sql  = """
            INSERT INTO yp_products
              (merchant_name, merchant_id, asin, product_name, category,
               price, commission, tracking_url, scraped_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE
              product_name=VALUES(product_name),
              price=VALUES(price),
              commission=VALUES(commission),
              tracking_url=VALUES(tracking_url),
              scraped_at=VALUES(scraped_at)
        """
        rows = [(p.get("merchant_name",""), p.get("merchant_id",""),
                 p.get("asin",""), p.get("product_name",""), p.get("category",""),
                 p.get("price",""), p.get("commission",""),
                 p.get("tracking_link",""), p.get("scraped_at",""))
                for p in products]
        cur.executemany(sql, rows)
        conn.commit()
        affected = cur.rowcount
        cur.close(); conn.close()
        return affected
    except Exception as e:
        log("DB", f"插入商品失败: {e}")
        return 0

def delete_products_from_db(mid: str, asins: set):
    """从 MySQL 删除已下架商品（标记删除）"""
    if not asins:
        return 0
    try:
        conn = get_db_conn()
        cur  = conn.cursor()
        placeholders = ",".join(["%s"] * len(asins))
        sql = f"DELETE FROM yp_products WHERE merchant_id=%s AND asin IN ({placeholders})"
        cur.execute(sql, [mid] + list(asins))
        conn.commit()
        deleted = cur.rowcount
        cur.close(); conn.close()
        return deleted
    except Exception as e:
        log("DB", f"删除商品失败: {e}")
        return 0

# ─── 拉取商户列表 ──────────────────────────────────────────────────────────
def fetch_merchants_from_api() -> list:
    """通过 YP API 拉取全量商户列表，用于发现新商户"""
    import requests
    token   = "7951dc7484fa9f9d"
    headers = {"token": token}
    url     = "https://www.yeahpromos.com/index/getadvert/getadvert"
    all_merchants = []
    page = 1
    while True:
        try:
            r    = requests.post(url, headers=headers,
                                 data={"site_id": SITE_ID, "page": page, "pagesize": 200},
                                 timeout=20)
            data = r.json()
            # 兼容两种响应格式
            items = (data.get("data", {}).get("data", {}).get("Data") or
                     data.get("data", {}).get("Data") or [])
            if not items:
                break
            for m in items:
                all_merchants.append({
                    "mid":    str(m.get("advert_id", "")),
                    "name":   m.get("advert_name", ""),
                    "status": m.get("join_status", ""),
                })
            if len(items) < 200:
                break
            page += 1
            time.sleep(0.2)
        except Exception as e:
            log("API", f"商户列表 page={page} 失败: {e}")
            break
    log("API", f"API 共获取商户: {len(all_merchants)} 个")
    return all_merchants

def load_local_merchants() -> list:
    """从本地 us_merchants_clean.json 加载商户列表（已筛选 APPROVED）"""
    path = SCRIPT_DIR / "output" / "us_merchants_clean.json"
    if path.exists():
        with open(str(path), 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data.get("approved_list", [])
    return []

# ─── Excel 工具 ────────────────────────────────────────────────────────────
def append_to_excel(products: list, sheet_name: str = "新增商品"):
    if not products:
        return
    import openpyxl
    from openpyxl.utils import get_column_letter
    path = Path(EXCEL_OUT)
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        wb = openpyxl.load_workbook(str(path))
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(COL_NAMES)
        ws.freeze_panes = "A2"
    for p in products:
        ws.append([p.get(k, "") for k in COL_KEYS])
    wb.save(str(path))

# ─── 文件解析（复用 download_only.py 的逻辑）──────────────────────────────
def parse_excel_file(filepath, merchant_id, merchant_name):
    try:
        fsize = os.path.getsize(filepath)
        if fsize < 100:
            return None
        with open(filepath, 'rb') as fh:
            magic = fh.read(2)
        if magic != b'PK':
            return None
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
        products = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            asin = str(row[0]).strip()
            if not asin or len(asin) != 10:
                continue
            products.append({
                "asin":          asin,
                "product_name":  str(row[1])[:200] if row[1] else "",
                "category":      str(row[2]) if row[2] else "",
                "commission":    str(row[3]) if row[3] else "",
                "price":         str(row[4]) if row[4] else "",
                "tracking_link": str(row[5]) if row[5] else "",
                "merchant_id":   merchant_id,
                "merchant_name": merchant_name,
                "scraped_at":    datetime.now().isoformat(),
            })
        wb.close()
        return products
    except Exception as e:
        log("EXCEL", f"解析失败: {e}")
        return None

def parse_html_fallback(page, mid, merchant_name):
    """HTML 分页兜底解析"""
    import requests as req_lib
    products = []
    cookies  = page.context.cookies()
    sess     = next((c for c in cookies if c["name"] == "PHPSESSID"), None)
    if not sess:
        html      = page.content()
        asin_list = re.findall(r'<div class="asin-code">([A-Z0-9]{10})</div>', html)
        link_list = [l.replace("&amp;", "&")
                     for l in re.findall(r"ClipboardJS\.copy\('([^']+)'\)", html)]
        for i, asin in enumerate(asin_list):
            products.append({"asin": asin, "product_name": "", "category": "",
                             "commission": "", "price": "",
                             "tracking_link": link_list[i] if i < len(link_list) else "",
                             "merchant_id": mid, "merchant_name": merchant_name,
                             "scraped_at": datetime.now().isoformat()})
        return products
    session = req_lib.Session()
    session.cookies.set("PHPSESSID", sess["value"])
    session.headers["User-Agent"] = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                                     "AppleWebKit/537.36")
    page_num = 1
    while True:
        try:
            url = (f"{BASE_URL}/index/offer/brand_detail"
                   f"?is_delete=0&advert_id={mid}&site_id={SITE_ID}&page={page_num}")
            r   = session.get(url, timeout=20)
            html      = r.text
            asin_list = re.findall(r'<div class="asin-code">([A-Z0-9]{10})</div>', html)
            link_list = [l.replace("&amp;", "&")
                         for l in re.findall(r"ClipboardJS\.copy\('([^']+)'\)", html)]
            if not asin_list:
                break
            for i, asin in enumerate(asin_list):
                products.append({"asin": asin, "product_name": "", "category": "",
                                 "commission": "", "price": "",
                                 "tracking_link": link_list[i] if i < len(link_list) else "",
                                 "merchant_id": mid, "merchant_name": merchant_name,
                                 "scraped_at": datetime.now().isoformat()})
            if len(asin_list) < 30:
                break
            page_num += 1
            time.sleep(0.3)
        except Exception as e:
            log("HTML", f"page={page_num} 解析失败: {e}")
            break
    return products

# ─── 核心：下载单个商户的商品列表 ─────────────────────────────────────────
def download_merchant_products(page, mid: str, name: str) -> list | None:
    """
    打开商户页面，下载 Excel 或 HTML 兜底。
    返回商品列表，None 表示失败（需重试）。
    """
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    url = f"{BASE_URL}/index/offer/brand_detail?is_delete=0&advert_id={mid}&site_id={SITE_ID}"
    try:
        page.goto(url, timeout=30000, wait_until="domcontentloaded")
        time.sleep(1.5)
        html = page.content()
        if "Login name cannot be empty" in html or (
                "login" in page.url.lower() and "yeahpromos" in page.url.lower()):
            log("MAIN", "⚠️  会话已过期，请在 Chrome 中重新登录，等待 60 秒...")
            time.sleep(60)
            page.goto(url, timeout=30000, wait_until="domcontentloaded")
            if "Login name cannot be empty" in page.content():
                log("MAIN", "仍未登录，中止")
                return None

        has_btn = page.evaluate("""
            () => {
                const links = Array.from(document.querySelectorAll('a'));
                return links.some(a => a.href && a.href.includes('export_advert_products'));
            }
        """)
        if not has_btn:
            return []   # 此商户无商品可下载

        save_path = os.path.join(DOWNLOAD_DIR, f"incr_{mid}.xlsx")
        try:
            with page.expect_download(timeout=20000) as dl_info:
                page.evaluate("""
                    () => {
                        const links = Array.from(document.querySelectorAll('a'));
                        const dl = links.find(a => a.href && a.href.includes('export_advert_products'));
                        if (dl) dl.click();
                    }
                """)
            dl_info.value.save_as(save_path)
            products = parse_excel_file(save_path, mid, name)
            try:
                os.remove(save_path)
            except Exception:
                pass
            if products is None:
                log("MAIN", "  Excel 无效，HTML 兜底...")
                products = parse_html_fallback(page, mid, name) or []
            return products
        except Exception as e_dl:
            log("MAIN", f"  下载异常: {type(e_dl).__name__}: {str(e_dl)[:80]}")
            products = parse_html_fallback(page, mid, name) or []
            return products

    except Exception as e:
        log("MAIN", f"  页面加载失败: {e}")
        return None

# ─── 主逻辑 ────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="YP 增量同步")
    parser.add_argument("--force-all", action="store_true",
                        help="强制重新对比所有商户（不跳过已同步的）")
    parser.add_argument("--new-only",  action="store_true",
                        help="只处理本次新发现的商户（跳过已有商户的商品差异对比）")
    parser.add_argument("--dry-run",   action="store_true",
                        help="只打印将要做的事，不实际修改任何数据")
    args = parser.parse_args()

    log("MAIN", "=" * 60)
    log("MAIN", f"YP 增量同步脚本 v1  {'[DRY RUN]' if args.dry_run else ''}")
    log("MAIN", "=" * 60)

    # ── Step 1: 加载商户列表（本地 + API 合并）──────────────────────────────
    local_merchants = load_local_merchants()
    local_mids      = {m["mid"] for m in local_merchants}
    log("MAIN", f"本地商户: {len(local_merchants)} 个")

    log("MAIN", "从 API 拉取最新商户列表...")
    api_merchants   = fetch_merchants_from_api()
    api_mids        = {m["mid"] for m in api_merchants}

    new_mids        = api_mids - local_mids           # 本次新增
    removed_mids    = local_mids - api_mids            # 已从平台移除（不强制删除，仅记录）

    log("MAIN", f"API 商户: {len(api_merchants)} 个")
    log("MAIN", f"✨ 新增商户: {len(new_mids)} 个")
    log("MAIN", f"⚠️  已移除商户: {len(removed_mids)} 个（仅记录，不删除数据）")

    # 合并商户列表：优先用 API 数据，兜底用本地
    mid_to_info = {m["mid"]: m for m in local_merchants}
    mid_to_info.update({m["mid"]: m for m in api_merchants})
    all_merchants   = list(mid_to_info.values())
    approved_mids   = {m["mid"] for m in all_merchants
                       if m.get("status", "").upper() == "APPROVED"}

    # ── Step 2: 加载同步状态 ────────────────────────────────────────────────
    sync_state = load_sync_state()

    # 决定本次需要处理哪些商户
    if args.new_only:
        to_process = [mid_to_info[m] for m in new_mids if m in mid_to_info]
        log("MAIN", f"--new-only 模式: 只处理 {len(to_process)} 个新商户")
    elif args.force_all:
        to_process = all_merchants
        log("MAIN", f"--force-all 模式: 重新处理全部 {len(to_process)} 个商户")
    else:
        # 默认：新商户 + 超过7天未同步的已有商户
        stale_threshold = 7 * 24 * 3600  # 7 天
        now_ts          = time.time()
        stale_mids      = set()
        for mid, snap in sync_state.items():
            last_sync = snap.get("last_sync", "")
            if last_sync:
                try:
                    last_ts = datetime.fromisoformat(last_sync).timestamp()
                    if now_ts - last_ts > stale_threshold:
                        stale_mids.add(mid)
                except Exception:
                    stale_mids.add(mid)
            else:
                stale_mids.add(mid)
        # 未同步过的商户也加进来
        never_synced = approved_mids - set(sync_state.keys())
        combined     = new_mids | stale_mids | never_synced
        to_process   = [mid_to_info[m] for m in combined if m in mid_to_info]
        log("MAIN",
            f"默认模式: 新商户={len(new_mids)} + 超期({stale_threshold//86400}天)={len(stale_mids)}"
            f" + 从未同步={len(never_synced)} → 共 {len(to_process)} 个商户待处理")

    if not to_process:
        log("MAIN", "✅ 所有商户均在同步期内，无需更新。退出。")
        return

    if args.dry_run:
        log("MAIN", "[DRY RUN] 将处理以下商户：")
        for m in to_process[:20]:
            log("MAIN", f"  {m['name']} (mid={m['mid']})")
        if len(to_process) > 20:
            log("MAIN", f"  ... 共 {len(to_process)} 个")
        return

    # ── Step 3: 连接 Chrome ─────────────────────────────────────────────────
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        log("MAIN", "未安装 playwright，无法运行")
        sys.exit(1)

    with sync_playwright() as pw:
        try:
            browser = pw.chromium.connect_over_cdp("http://localhost:9222")
            log("MAIN", "✅ 成功连接调试 Chrome")
        except Exception as e:
            log("MAIN", f"❌ 无法连接 Chrome（9222端口）: {e}")
            sys.exit(1)

        ctx  = browser.contexts[0]
        page = ctx.pages[0] if ctx.pages else ctx.new_page()

        # ── Step 4: 逐商户增量对比 ──────────────────────────────────────────
        stats = {"new_merchants": 0, "added_products": 0,
                 "removed_products": 0, "unchanged": 0, "errors": 0}
        new_excel_rows = []
        start_time     = time.time()

        for idx, merchant in enumerate(to_process):
            mid  = merchant["mid"]
            name = merchant.get("name", mid)
            is_new = mid in new_mids
            elapsed = int(time.time() - start_time)
            log("MAIN",
                f"[{idx+1}/{len(to_process)}] {'🆕 ' if is_new else ''}{name} "
                f"(mid={mid}) | {elapsed//60}m{elapsed%60}s")

            products = download_merchant_products(page, mid, name)
            if products is None:
                log("MAIN", "  ⚠️  下载失败，跳过")
                stats["errors"] += 1
                continue

            current_asins = {p["asin"] for p in products}

            if is_new:
                # 新商户：全量插入
                if products:
                    n = insert_products_to_db(products)
                    log("MAIN", f"  🆕 新商户，插入 {len(products)} 条（DB影响行={n}）")
                    new_excel_rows.extend(products)
                    stats["new_merchants"] += 1
                    stats["added_products"] += len(products)
                else:
                    log("MAIN", "  新商户，暂无商品")
            else:
                # 旧商户：差异对比
                db_asins = get_existing_asins_from_db(mid)
                added    = current_asins - db_asins
                removed  = db_asins - current_asins

                if added:
                    add_products = [p for p in products if p["asin"] in added]
                    n = insert_products_to_db(add_products)
                    log("MAIN", f"  ➕ 新增商品: {len(added)} 条")
                    new_excel_rows.extend(add_products)
                    stats["added_products"] += len(added)

                if removed:
                    n = delete_products_from_db(mid, removed)
                    log("MAIN", f"  ➖ 下架商品: {len(removed)} 条（已从DB删除）")
                    stats["removed_products"] += len(removed)

                if not added and not removed:
                    log("MAIN", f"  ✓  无变化 ({len(current_asins)} 条)")
                    stats["unchanged"] += 1

            # 更新快照
            sync_state[mid] = {
                "name":       name,
                "count":      len(products),
                "asins":      list(current_asins),
                "last_sync":  datetime.now().isoformat(),
            }
            save_sync_state(sync_state)
            time.sleep(0.5)

        # ── Step 5: 写入 Excel 增量文件 ─────────────────────────────────────
        if new_excel_rows:
            append_to_excel(new_excel_rows, sheet_name="新增商品")
            log("MAIN", f"📄 Excel 写入: {len(new_excel_rows)} 条 → {EXCEL_OUT}")

        # ── Step 6: 汇总报告 ────────────────────────────────────────────────
        elapsed_total = int(time.time() - start_time)
        log("MAIN", "=" * 60)
        log("MAIN", "同步完成报告")
        log("MAIN", f"  总耗时:    {elapsed_total//60}m{elapsed_total%60}s")
        log("MAIN", f"  处理商户:  {len(to_process)} 个")
        log("MAIN", f"  新商户:    {stats['new_merchants']} 个")
        log("MAIN", f"  新增商品:  {stats['added_products']} 条")
        log("MAIN", f"  下架商品:  {stats['removed_products']} 条")
        log("MAIN", f"  无变化:    {stats['unchanged']} 个")
        log("MAIN", f"  失败商户:  {stats['errors']} 个")
        log("MAIN", "=" * 60)

        # ── Step 7: 自动同步缓存表 ──────────────────────────────────────────
        if stats['added_products'] > 0 and not args.dry_run:
            try:
                log("MAIN", "正在同步 yp_us_products 缓存表...")
                from build_us_cache import incremental_refresh
                incremental_refresh()
            except Exception as e:
                log("MAIN", f"缓存表同步失败（不影响同步结果）: {e}")

if __name__ == "__main__":
    main()
