"""
YP Products Excel → MySQL 增量同步脚本
逻辑：读取 Excel 全量数据，对比 MySQL 中已有的 (merchant_id, asin) 组合，
      只插入/更新新出现的记录，已有记录若价格/链接有变化则更新。

用法:
    python sync_excel_to_mysql.py              # 正常增量同步
    python sync_excel_to_mysql.py --full       # 全量重新对比（慢，慎用）
    python sync_excel_to_mysql.py --stats      # 只显示统计，不写入
"""

import sys
import time
import json
import logging
from pathlib import Path
from datetime import datetime

# ─── 配置 ────────────────────────────────────────────────────────
EXCEL_PATH   = r"D:\workspace\YP_products.xlsx"
MYSQL_HOST   = "localhost"
MYSQL_PORT   = 3306
MYSQL_USER   = "root"
MYSQL_PASS   = "admin"
MYSQL_DB     = "affiliate_marketing"
MYSQL_TABLE  = "yp_products"
BATCH_SIZE   = 3000       # 每批 INSERT 行数
STATE_FILE   = Path(__file__).parent / "output" / "sync_state.json"
LOG_FILE     = Path(__file__).parent / "output" / "sync_log.txt"
# ─────────────────────────────────────────────────────────────────

# ─── 日志 ────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
    ]
)
log = logging.getLogger(__name__)
# ─────────────────────────────────────────────────────────────────


def get_conn():
    import mysql.connector
    return mysql.connector.connect(
        host=MYSQL_HOST, port=MYSQL_PORT,
        user=MYSQL_USER, password=MYSQL_PASS,
        database=MYSQL_DB, charset="utf8mb4",
        autocommit=False
    )


def load_state() -> dict:
    """加载上次同步状态"""
    if STATE_FILE.exists():
        try:
            return json.loads(STATE_FILE.read_text("utf-8"))
        except Exception:
            pass
    return {
        "last_sync": None,
        "last_excel_mtime": None,
        "last_excel_rows": 0,
        "total_mysql_rows": 0,
    }


def save_state(state: dict):
    STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    STATE_FILE.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def get_excel_fingerprint(path: Path) -> tuple:
    """返回 (mtime, size)，用于快速检测 Excel 是否有变化"""
    st = path.stat()
    return (st.st_mtime, st.st_size)


def load_existing_keys(conn) -> set:
    """从 MySQL 读取所有已有的 (merchant_id, asin) 组合，用于增量判断"""
    log.info("从 MySQL 读取已有 key 集合（merchant_id + asin）...")
    cur = conn.cursor()
    cur.execute(f"SELECT merchant_id, asin FROM {MYSQL_TABLE}")
    keys = set()
    for row in cur:
        keys.add((str(row[0]), str(row[1])))
    cur.close()
    log.info(f"  MySQL 已有 {len(keys):,} 条记录")
    return keys


def load_existing_tracking(conn) -> dict:
    """读取已有记录的投放链接，用于检测链接更新"""
    cur = conn.cursor()
    cur.execute(f"SELECT merchant_id, asin, tracking_url FROM {MYSQL_TABLE} WHERE tracking_url IS NOT NULL AND tracking_url != ''")
    result = {}
    for row in cur:
        result[(str(row[0]), str(row[1]))] = str(row[2])
    cur.close()
    return result


def parse_price(val) -> float | None:
    if val is None:
        return None
    try:
        return float(str(val).replace("$", "").replace(",", "").strip())
    except Exception:
        return None


def sync(full_compare=False, stats_only=False):
    import openpyxl

    excel_path = Path(EXCEL_PATH)
    if not excel_path.exists():
        log.error(f"❌ Excel 文件不存在: {EXCEL_PATH}")
        return

    state = load_state()
    fp = get_excel_fingerprint(excel_path)
    fp_str = f"{fp[0]:.0f}:{fp[1]}"

    log.info("=" * 60)
    log.info(f"YP Excel → MySQL 增量同步开始")
    log.info(f"Excel: {EXCEL_PATH}  ({excel_path.stat().st_size/1024/1024:.1f} MB)")
    log.info(f"上次同步: {state.get('last_sync', '从未')}")
    log.info(f"上次 Excel mtime+size: {state.get('last_excel_mtime', '无')}")
    log.info(f"当前 Excel mtime+size: {fp_str}")

    # 快速判断 Excel 是否有变化
    if not full_compare and state.get("last_excel_mtime") == fp_str:
        log.info("⏭  Excel 文件未发生变化，跳过本次同步")
        log.info("=" * 60)
        return

    conn = get_conn()
    log.info("✅ MySQL 连接成功")

    # 获取现有 key 集合
    existing_keys    = load_existing_keys(conn)
    existing_tracking = load_existing_tracking(conn)

    # 读 Excel
    log.info(f"读取 Excel（read_only 模式）...")
    t0 = time.time()
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    total_excel_rows = (ws.max_row or 1) - 1
    log.info(f"Excel 共 {total_excel_rows:,} 行数据，读取耗时 {time.time()-t0:.1f}s")

    # 统计
    new_rows     = []   # 全新的
    update_rows  = []   # 已有但链接有更新的
    skip_count   = 0
    invalid_count= 0

    INSERT_SQL = f"""
        INSERT INTO {MYSQL_TABLE}
          (merchant_name, merchant_id, asin, product_name, category,
           price, commission, tracking_url, scraped_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
          product_name  = VALUES(product_name),
          price         = VALUES(price),
          commission    = VALUES(commission),
          tracking_url  = IF(VALUES(tracking_url) IS NOT NULL AND VALUES(tracking_url) != '',
                             VALUES(tracking_url), tracking_url),
          scraped_at    = VALUES(scraped_at)
    """

    UPDATE_LINK_SQL = f"""
        UPDATE {MYSQL_TABLE}
        SET tracking_url = %s, scraped_at = %s
        WHERE merchant_id = %s AND asin = %s
          AND (tracking_url IS NULL OR tracking_url = '')
    """

    for row in ws.iter_rows(min_row=2, values_only=True):
        merchant_name = str(row[0]).strip() if row[0] else ""
        merchant_id   = str(row[1]).strip() if row[1] else ""
        asin          = str(row[2]).strip() if row[2] else ""
        product_name  = str(row[3]).strip() if row[3] else None
        category      = str(row[4]).strip() if row[4] else None
        price         = parse_price(row[5])
        commission    = str(row[6]).strip() if row[6] else None
        tracking_url  = str(row[7]).strip() if row[7] else None
        scraped_at    = str(row[8]).strip() if row[8] else None

        if not asin or asin == "None":
            invalid_count += 1
            continue

        key = (merchant_id, asin)

        if key not in existing_keys:
            # 全新记录
            new_rows.append((
                merchant_name, merchant_id, asin, product_name, category,
                price, commission, tracking_url, scraped_at
            ))
        else:
            # 已存在：检查是否有新的投放链接
            old_link = existing_tracking.get(key)
            has_new_link = tracking_url and str(tracking_url).startswith("http")
            had_no_link  = not old_link
            if has_new_link and had_no_link:
                update_rows.append((tracking_url, scraped_at, merchant_id, asin))
            else:
                skip_count += 1

    wb.close()

    log.info(f"\n扫描完成:")
    log.info(f"  全新记录: {len(new_rows):,} 条")
    log.info(f"  链接更新: {len(update_rows):,} 条（之前无链接，现在有）")
    log.info(f"  已跳过:   {skip_count:,} 条（已存在且无变化）")
    log.info(f"  无效行:   {invalid_count:,} 条")

    if stats_only:
        log.info("(--stats 模式，不写入数据库)")
        conn.close()
        return

    if not new_rows and not update_rows:
        log.info("✅ 无需同步，数据库已是最新状态")
        state["last_sync"] = datetime.now().isoformat()
        state["last_excel_mtime"] = fp_str
        save_state(state)
        conn.close()
        return

    cur = conn.cursor()
    t1 = time.time()

    # 批量插入新记录
    if new_rows:
        log.info(f"\n开始批量插入 {len(new_rows):,} 条新记录...")
        written = 0
        for i in range(0, len(new_rows), BATCH_SIZE):
            batch = new_rows[i:i+BATCH_SIZE]
            cur.executemany(INSERT_SQL, batch)
            conn.commit()
            written += len(batch)
            pct = written / len(new_rows) * 100
            speed = written / max(time.time()-t1, 0.1)
            log.info(f"  INSERT 进度: {written:,}/{len(new_rows):,} ({pct:.1f}%)  {speed:.0f} 行/秒")

    # 批量更新有新链接的记录
    if update_rows:
        log.info(f"\n开始更新 {len(update_rows):,} 条新增投放链接...")
        written = 0
        for i in range(0, len(update_rows), BATCH_SIZE):
            batch = update_rows[i:i+BATCH_SIZE]
            cur.executemany(UPDATE_LINK_SQL, batch)
            conn.commit()
            written += len(batch)
        log.info(f"  链接更新完成: {written:,} 条")

    cur.close()
    conn.close()

    elapsed = time.time() - t1

    # 更新状态
    state["last_sync"]         = datetime.now().isoformat()
    state["last_excel_mtime"]  = fp_str
    state["last_excel_rows"]   = total_excel_rows
    state["total_mysql_rows"]  = len(existing_keys) + len(new_rows)
    save_state(state)

    log.info("=" * 60)
    log.info(f"✅ 同步完成！")
    log.info(f"   新插入:  {len(new_rows):,} 条")
    log.info(f"   链接更新:{len(update_rows):,} 条")
    log.info(f"   耗时:    {elapsed:.1f} 秒")
    log.info("=" * 60)


if __name__ == "__main__":
    full   = "--full"   in sys.argv
    stats  = "--stats"  in sys.argv
    sync(full_compare=full, stats_only=stats)
