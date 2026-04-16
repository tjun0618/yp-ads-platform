# -*- coding: utf-8 -*-
"""
分析 export_top_merchants 返回的 Excel 文件
"""
import requests
import os

TOKEN = '7951dc7484fa9f9d'
SITE_ID = '12002'
HEADERS = {'token': TOKEN}
COOKIE = {'PHPSESSID': '932a965dc80f3c5bc7fe2226771950fc'}

BASE = 'https://yeahpromos.com'

print("=" * 80)
print("分析 Export API")
print("=" * 80)

# 1. Download the export file
url = f"{BASE}/index/advert/export_top_merchants"
print(f"\n[1] Downloading from {url}")

resp = requests.get(url, headers=HEADERS, cookies=COOKIE, timeout=30)
print(f"Status: {resp.status_code}")
print(f"Content-Type: {resp.headers.get('Content-Type', 'N/A')}")
print(f"Content-Length: {len(resp.content)} bytes")

# Save the file
excel_path = 'output/export_top_merchants.xlsx'
with open(excel_path, 'wb') as f:
    f.write(resp.content)
print(f"Saved to {excel_path}")

# 2. Try to read it
print("\n[2] Reading Excel file...")
try:
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    
    print(f"Sheets: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n  Sheet: {sheet_name}")
        print(f"  Rows: {ws.max_row}, Cols: {ws.max_column}")
        
        # Print headers
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        print(f"  Headers: {headers}")
        
        # Print first 5 rows
        print(f"\n  First 5 rows:")
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True)):
            print(f"    Row {i+1}: {row}")
    
    wb.close()
    print(f"\n  Total data rows: {ws.max_row - 1}")
    
except Exception as e:
    print(f"  Error reading Excel: {e}")

# 3. Check what the HTML responses contain
print("\n[3] Checking HTML responses...")
html_files = [
    'output/page_Brands_List.html',
    'output/page_Brand_Detail_Page.html',
]

for html_file in html_files:
    if os.path.exists(html_file):
        with open(html_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        print(f"\n  {html_file}:")
        print(f"    Length: {len(content)}")
        
        # Check if it's an error page
        if 'login' in content.lower() or 'sign in' in content.lower():
            print(f"    WARNING: Contains 'login' - might be redirect to login page")
        
        # Look for title
        import re
        title = re.search(r'<title>([^<]+)</title>', content, re.IGNORECASE)
        if title:
            print(f"    Title: {title.group(1)}")
        
        # Check for specific content
        if 'brand_detail' in html_file:
            # Look for product data
            products = re.findall(r'data-product-id=["\']([^"\']+)["\']', content)
            print(f"    Product IDs found: {len(products)}")
            
            # Look for ASINs
            asins = re.findall(r'[A-Z0-9]{10}', content)
            unique_asins = set(asins)
            print(f"    ASIN-like strings: {len(unique_asins)}")

print("\n" + "=" * 80)
