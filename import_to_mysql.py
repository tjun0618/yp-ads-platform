"""
YP Products Excel → MySQL 导入脚本
将 D:\\workspace\\YP_products.xlsx 批量导入到 affiliate_marketing.yp_products 表
"""

import sys
import os
import time
from pathlib import Path
from datetime import datetime

# ─── 配置 ───────────────────────────────────────────────────────
EXCEL_PATH  = r"D:\workspace\YP_products.xlsx"
MYSQL_HOST  = "localhost"
MYSQL_PORT  = 3306
MYSQL_USER  = "root"
MYSQL_PASS  = "admin"
MYSQL_DB    = "affiliate_marketing"
MYSQL_TABLE = "yp_products"
BATCH_SIZE  = 2000   # 每批插入行数，平衡速度与内存
# ────────────────────────────────────────────────────────────────

LOG_FILE = Path(__file__).parent / "output" / "mysql_import_log.txt"

def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line, flush=True)
    try:
        with open(LOG_FILE, 'a', encoding='utf-8') as f:
            f.write(line + "\n")
    except Exception:
        pass

def check_deps():
    """检查并安装依赖"""
    missing = []
    try:
        import mysql.connector
    except ImportError:
        missing.append("mysql-connector-python")
    try:
        import openpyxl
    except ImportError:
        missing.append("openpyxl")

    if missing:
        log(f"安装依赖: {missing}")
        import subprocess
        subprocess.run([sys.executable, "-m", "pip", "install"] + missing,
                       check=True, capture_output=True)
        log("依赖安装完成")

def get_connection():
    import mysql.connector
    return mysql.connector.connect(
        host=MYSQL_HOST,
        port=MYSQL_PORT,
        user=MYSQL_USER,
        password=MYSQL_PASS,
        database=MYSQL_DB,
        charset="utf8mb4",
        autocommit=False
    )

def truncate_or_skip(conn, force_truncate=False):
    """检查表是否已有数据"""
    cur = conn.cursor()
    cur.execute(f"SELECT COUNT(*) FROM {MYSQL_TABLE}")
    count = cur.fetchone()[0]
    cur.close()
    if count > 0:
        if force_truncate:
            log(f"表已有 {count:,} 条，执行 TRUNCATE 清空后重新导入...")
            cur2 = conn.cursor()
            cur2.execute(f"TRUNCATE TABLE {MYSQL_TABLE}")
            conn.commit()
            cur2.close()
            return 0
        else:
            log(f"表已有 {count:,} 条数据，跳过已导入部分，继续追加新数据")
            return count
    return 0

def import_excel_to_mysql(force_truncate=False):
    check_deps()
    import openpyxl

    log("=" * 55)
    log("YP Products Excel → MySQL 导入开始")
    log(f"源文件: {EXCEL_PATH}")
    log(f"目标: {MYSQL_DB}.{MYSQL_TABLE} @ {MYSQL_HOST}")
    log("=" * 55)

    # 检查文件
    excel_path = Path(EXCEL_PATH)
    if not excel_path.exists():
        log(f"❌ 文件不存在: {EXCEL_PATH}")
        return

    file_size_mb = excel_path.stat().st_size / 1024 / 1024
    log(f"文件大小: {file_size_mb:.1f} MB，开始读取（请稍候）...")

    # 用 read_only 模式读，节省内存
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    total_rows = ws.max_row - 1  # 减去表头
    log(f"总数据行: {total_rows:,}")

    # 连接 MySQL
    conn = get_connection()
    log("✅ MySQL 连接成功")

    # 处理已有数据
    existing = truncate_or_skip(conn, force_truncate)

    # 表头列映射（Excel列顺序）
    # 商户名称, 商户ID, ASIN, 商品名称, 类别, 价格(USD), 佣金率, 投放链接, 采集时间
    INSERT_SQL = f"""
        INSERT INTO {MYSQL_TABLE}
          (merchant_name, merchant_id, asin, product_name, category,
           price, commission, tracking_url, scraped_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
          product_name = VALUES(product_name),
          price        = VALUES(price),
          commission   = VALUES(commission),
          tracking_url = VALUES(tracking_url),
          scraped_at   = VALUES(scraped_at)
    """

    cur = conn.cursor()
    batch = []
    total_inserted = 0
    total_skipped  = 0
    start_time = time.time()
    row_num = 0

    log("开始逐行读取并批量写入 MySQL...")

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1

        # 解析每列
        merchant_name = str(row[0]).strip() if row[0] else ""
        merchant_id   = str(row[1]).strip() if row[1] else ""
        asin          = str(row[2]).strip() if row[2] else ""
        product_name  = str(row[3]).strip() if row[3] else None
        category      = str(row[4]).strip() if row[4] else None
        price_raw     = row[5]
        commission    = str(row[6]).strip() if row[6] else None
        tracking_url  = str(row[7]).strip() if row[7] else None
        scraped_at    = str(row[8]).strip() if row[8] else None

        # 跳过无效行
        if not asin or asin == "None":
            total_skipped += 1
            continue

        # 解析价格
        price = None
        if price_raw:
            try:
                price = float(str(price_raw).replace('$', '').replace(',', '').strip())
            except Exception:
                price = None

        batch.append((
            merchant_name, merchant_id, asin, product_name, category,
            price, commission, tracking_url, scraped_at
        ))

        # 批量提交
        if len(batch) >= BATCH_SIZE:
            cur.executemany(INSERT_SQL, batch)
            conn.commit()
            total_inserted += len(batch)
            batch = []

            elapsed = time.time() - start_time
            speed = total_inserted / elapsed if elapsed > 0 else 0
            remain = (total_rows - row_num) / speed if speed > 0 else 0
            log(f"  进度: {row_num:,}/{total_rows:,} ({row_num/total_rows*100:.1f}%) | "
                f"已写入: {total_inserted:,} | 速度: {speed:.0f} 行/秒 | "
                f"剩余: {remain/60:.1f} 分钟")

    # 提交最后一批
    if batch:
        cur.executemany(INSERT_SQL, batch)
        conn.commit()
        total_inserted += len(batch)

    cur.close()
    conn.close()
    wb.close()

    elapsed = time.time() - start_time
    log("=" * 55)
    log(f"✅ 导入完成！")
    log(f"   总数据行:  {total_rows:,}")
    log(f"   成功写入:  {total_inserted:,}")
    log(f"   跳过无效:  {total_skipped:,}")
    log(f"   总耗时:    {elapsed/60:.1f} 分钟")
    log(f"   平均速度:  {total_inserted/elapsed:.0f} 行/秒")
    log("=" * 55)

if __name__ == "__main__":
    # 命令行加 --truncate 参数则清空重导
    force = "--truncate" in sys.argv
    import_excel_to_mysql(force_truncate=force)
