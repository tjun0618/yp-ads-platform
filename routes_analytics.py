# routes_analytics.py - 投放优化、质量评分、竞品分析、YP同步
# 从 ads_manager.py 行 5890-8862 提取
import json, os, sys, io, csv, uuid, re, time as _time, requests, threading, subprocess, datetime
from pathlib import Path
from flask import (
    Blueprint,
    render_template_string,
    jsonify,
    request,
    redirect,
    url_for,
    send_file,
    Response,
    current_app,
)
from urllib.parse import quote
from app_config import (
    DB,
    BASE_DIR,
    OUTPUT_DIR,
    YP_SYNC_SCRIPT,
    YP_SYNC_STATE,
    YP_SYNC_LOG,
)
from db import (
    get_db,
    _db,
    _cached_count,
    _count_cache,
    _parse_report_column,
    _clean_numeric_value,
)
from templates_shared import BASE_CSS, NAV_HTML, _BASE_STYLE_DARK, _SCRAPE_TOPNAV

# 创建 Blueprint
bp = Blueprint("analytics", __name__)


# ═══════════════════════════════════════════════════════════════════════════
# 投放优化 - 商户数据 API
# ═══════════════════════════════════════════════════════════════════════════


@bp.route("/api/optimize/merchants")
def api_optimize_merchants():
    """获取商户列表（用于下拉选择）"""
    try:
        conn = get_db()
        cur = conn.cursor(dictionary=True)

        cur.execute(
            """
            SELECT DISTINCT merchant_id, merchant_name
            FROM yp_merchants
            WHERE merchant_name IS NOT NULL AND merchant_name != ''
            ORDER BY merchant_name
            """
        )
        merchants = cur.fetchall()
        conn.close()

        return jsonify({"success": True, "merchants": merchants})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@bp.route("/api/optimize/yp_data/<merchant_id>", methods=["GET", "POST"])
def api_optimize_yp_data(merchant_id):
    """获取或拉取 YP 品牌数据"""
    try:
        from datetime import datetime
        import subprocess
        import sys

        force_refresh = request.args.get("force", "false").lower() == "true"

        # 日期范围：优先使用传入参数，否则默认 2026-01-01 到当天
        start_date = request.args.get("start_date", "")
        end_date = request.args.get("end_date", "")

        if not start_date or not end_date:
            today = datetime.now()
            start_date = "2026-01-01"
            end_date = today.strftime("%Y-%m-%d")

        conn = get_db()
        cur = conn.cursor(dictionary=True)

        # 获取商户名称
        cur.execute(
            "SELECT merchant_name FROM yp_merchants WHERE merchant_id = %s LIMIT 1",
            (merchant_id,),
        )
        merchant = cur.fetchone()
        merchant_name = merchant["merchant_name"] if merchant else f"商户{merchant_id}"

        # 强制刷新：调用采集脚本
        if force_refresh and request.method == "POST":
            script_path = BASE_DIR / "yp_report_collector.py"
            if script_path.exists():
                print(
                    f"[Optimize] 正在从YP平台获取 {merchant_name} 的数据 ({start_date} ~ {end_date})..."
                )

                # 异步执行采集脚本
                subprocess.Popen(
                    [
                        sys.executable,
                        str(script_path),
                        "--type",
                        "all",
                        "--start-date",
                        start_date,
                        "--end-date",
                        end_date,
                    ],
                    cwd=str(BASE_DIR),
                )

                conn.close()
                return jsonify(
                    {
                        "success": True,
                        "collecting": True,
                        "message": f"正在从YP平台获取数据 ({start_date} ~ {end_date})，请稍后刷新查看",
                    }
                )
            else:
                conn.close()
                return jsonify({"success": False, "error": "采集脚本不存在"})

        # 查询商户报表数据
        cur.execute(
            """
            SELECT merchant_id, merchant_name, report_date, clicks, detail_views,
                   add_to_carts, purchases, amount, commission
            FROM yp_merchant_report
            WHERE merchant_id = %s AND report_date = %s
            LIMIT 1
            """,
            (merchant_id, end_date),
        )
        merchant = cur.fetchone()
        merchant_name = merchant["merchant_name"] if merchant else f"商户{merchant_id}"

        # 强制刷新：调用采集脚本
        if force_refresh and request.method == "POST":
            script_path = BASE_DIR / "yp_report_collector.py"
            if script_path.exists():
                print(f"[Optimize] 正在从YP平台获取 {merchant_name} 的数据...")

                # 异步执行采集脚本
                subprocess.Popen(
                    [
                        sys.executable,
                        str(script_path),
                        "--type",
                        "all",
                        "--start-date",
                        start_date,
                        "--end-date",
                        end_date,
                    ],
                    cwd=str(BASE_DIR),
                )

                conn.close()
                return jsonify(
                    {
                        "success": True,
                        "collecting": True,
                        "message": "正在从YP平台获取数据，请稍后刷新查看",
                    }
                )
            else:
                conn.close()
                return jsonify({"success": False, "error": "采集脚本不存在"})

        # 查询商户报表数据
        cur.execute(
            """
            SELECT merchant_id, merchant_name, report_date, clicks, detail_views,
                   add_to_carts, purchases, amount, commission
            FROM yp_merchant_report
            WHERE merchant_id = %s
            ORDER BY report_date DESC
            LIMIT 1
            """,
            (merchant_id,),
        )
        merchant_report = cur.fetchone()

        # 查询商品报表数据（关联商品名称）
        cur.execute(
            """
            SELECT p.product_id, p.asin, COALESCE(pr.product_name, '') as product_name,
                   p.clicks, p.detail_views, p.add_to_carts, p.purchases, p.amount, p.commission
            FROM yp_product_report p
            LEFT JOIN yp_us_products pr ON p.asin = CAST(pr.asin AS CHAR CHARACTER SET utf8mb4)
            WHERE p.merchant_id = %s
            ORDER BY p.commission DESC, p.clicks DESC
            LIMIT 50
            """,
            (merchant_id,),
        )
        product_reports = cur.fetchall()
        conn.close()

        if merchant_report:
            return jsonify(
                {
                    "success": True,
                    "data": {
                        "merchant_id": merchant_report["merchant_id"],
                        "merchant_name": merchant_report["merchant_name"],
                        "report_date": str(merchant_report["report_date"]),
                        "clicks": merchant_report["clicks"] or 0,
                        "detail_views": merchant_report["detail_views"] or 0,
                        "add_to_carts": merchant_report["add_to_carts"] or 0,
                        "purchases": merchant_report["purchases"] or 0,
                        "amount": float(merchant_report["amount"] or 0),
                        "commission": float(merchant_report["commission"] or 0),
                    },
                    "products": [
                        {
                            "product_id": p["product_id"],
                            "asin": p["asin"],
                            "product_name": p["product_name"] or "",
                            "clicks": p["clicks"] or 0,
                            "detail_views": p["detail_views"] or 0,
                            "add_to_carts": p["add_to_carts"] or 0,
                            "purchases": p["purchases"] or 0,
                            "amount": float(p["amount"] or 0),
                            "commission": float(p["commission"] or 0),
                        }
                        for p in product_reports
                    ],
                }
            )
        else:
            return jsonify(
                {
                    "success": True,
                    "data": {
                        "merchant_id": merchant_id,
                        "merchant_name": merchant_name,
                        "report_date": None,
                        "clicks": 0,
                        "detail_views": 0,
                        "add_to_carts": 0,
                        "purchases": 0,
                        "amount": 0.0,
                        "commission": 0.0,
                    },
                    "products": [],
                }
            )

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@bp.route("/api/optimize/save_ads_cost", methods=["POST"])
def api_optimize_save_ads_cost():
    """保存商户的广告花费（人民币）"""
    try:
        data = request.get_json()
        merchant_id = data.get("merchant_id")
        ads_cost_cny = data.get("ads_cost_cny", 0)

        if not merchant_id:
            return jsonify({"success": False, "error": "缺少商户ID"})

        conn = get_db()
        cur = conn.cursor()

        # 更新该商户最新的记录的广告花费
        cur.execute(
            """
            UPDATE yp_merchant_report
            SET ads_cost_cny = %s
            WHERE merchant_id = %s
            ORDER BY report_date DESC
            LIMIT 1
            """,
            (ads_cost_cny, merchant_id),
        )

        # 如果没有更新任何记录，插入一条新记录
        if cur.rowcount == 0:
            cur.execute(
                """
                INSERT INTO yp_merchant_report (merchant_id, report_date, ads_cost_cny)
                VALUES (%s, CURDATE(), %s)
                """,
                (merchant_id, ads_cost_cny),
            )

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({"success": True})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@bp.route("/api/optimize/get_ads_cost/<merchant_id>", methods=["GET"])
def api_optimize_get_ads_cost(merchant_id):
    """获取商户的广告花费（人民币）"""
    try:
        conn = get_db()
        cur = conn.cursor(dictionary=True)

        cur.execute(
            """
            SELECT ads_cost_cny FROM yp_merchant_report
            WHERE merchant_id = %s AND ads_cost_cny > 0
            ORDER BY report_date DESC, updated_at DESC
            LIMIT 1
            """,
            (merchant_id,),
        )
        result = cur.fetchone()
        cur.close()
        conn.close()

        if result and result["ads_cost_cny"]:
            return jsonify(
                {"success": True, "ads_cost_cny": float(result["ads_cost_cny"])}
            )
        else:
            return jsonify({"success": True, "ads_cost_cny": None})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@bp.route("/api/optimize/total_ads_cost", methods=["GET"])
def api_optimize_total_ads_cost():
    """获取所有商户的 Google Ads 总花费和总佣金（人民币）"""
    try:
        conn = get_db()
        cur = conn.cursor(dictionary=True)

        # 获取每个商户最新的广告花费
        cur.execute(
            """
            SELECT SUM(ads_cost_cny) as total_cost
            FROM (
                SELECT merchant_id, ads_cost_cny
                FROM yp_merchant_report r1
                WHERE ads_cost_cny > 0
                AND id = (
                    SELECT id FROM yp_merchant_report r2
                    WHERE r2.merchant_id = r1.merchant_id AND ads_cost_cny > 0
                    ORDER BY report_date DESC, updated_at DESC
                    LIMIT 1
                )
            ) as latest_costs
            """
        )
        cost_result = cur.fetchone()

        # 获取每个商户最新的佣金
        cur.execute(
            """
            SELECT SUM(commission) as total_commission_usd
            FROM (
                SELECT merchant_id, commission
                FROM yp_merchant_report r1
                WHERE commission > 0
                AND id = (
                    SELECT id FROM yp_merchant_report r2
                    WHERE r2.merchant_id = r1.merchant_id
                    ORDER BY report_date DESC, updated_at DESC
                    LIMIT 1
                )
            ) as latest_commissions
            """
        )
        commission_result = cur.fetchone()

        cur.close()
        conn.close()

        total_cost = (
            float(cost_result["total_cost"])
            if cost_result and cost_result["total_cost"]
            else 0
        )
        total_commission_usd = (
            float(commission_result["total_commission_usd"])
            if commission_result and commission_result["total_commission_usd"]
            else 0
        )

        # 佣金换算成人民币 (汇率 7.2)
        EXCHANGE_RATE = 7.2
        total_commission_cny = total_commission_usd * EXCHANGE_RATE

        # 利润 = 总佣金 - 总花费
        total_profit = total_commission_cny - total_cost

        return jsonify(
            {
                "success": True,
                "total_ads_cost": total_cost,
                "total_commission": total_commission_cny,
                "total_profit": total_profit,
            }
        )

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@bp.route("/api/optimize/download_excel/<merchant_id>", methods=["GET"])
def api_optimize_download_excel(merchant_id):
    """下载商户销售数据 Excel"""
    try:
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # 获取日期参数
        start_date = request.args.get("start_date", "")
        end_date = request.args.get("end_date", "")

        if not start_date or not end_date:
            today = datetime.now()
            start_date = today.strftime("%Y-%m-01")
            end_date = today.strftime("%Y-%m-%d")

        conn = get_db()
        cur = conn.cursor(dictionary=True)

        # 获取商户名称
        cur.execute(
            "SELECT merchant_name FROM yp_merchants WHERE merchant_id = %s LIMIT 1",
            (merchant_id,),
        )
        merchant = cur.fetchone()
        merchant_name = merchant["merchant_name"] if merchant else f"商户{merchant_id}"

        # 查询商户报表数据
        cur.execute(
            """
            SELECT merchant_id, merchant_name, report_date, clicks, detail_views,
                   add_to_carts, purchases, amount, commission
            FROM yp_merchant_report
            WHERE merchant_id = %s
            ORDER BY report_date DESC
            LIMIT 1
            """,
            (merchant_id,),
        )
        merchant_report = cur.fetchone()

        # 查询商品报表数据
        cur.execute(
            """
            SELECT p.product_id, p.asin, COALESCE(pr.product_name, '') as product_name,
                   p.clicks, p.detail_views, p.add_to_carts, p.purchases, p.amount, p.commission
            FROM yp_product_report p
            LEFT JOIN yp_us_products pr ON p.asin = CAST(pr.asin AS CHAR CHARACTER SET utf8mb4)
            WHERE p.merchant_id = %s
            ORDER BY p.commission DESC, p.clicks DESC
            """,
            (merchant_id,),
        )
        product_reports = cur.fetchall()
        conn.close()

        # 创建 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "销售数据"

        # 样式定义
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_alignment = Alignment(horizontal="center", vertical="center")
        number_alignment = Alignment(horizontal="right", vertical="center")
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        highlight_fill = PatternFill("solid", fgColor="E2EFDA")

        # 标题行
        ws.merge_cells("A1:H1")
        ws["A1"] = f"{merchant_name} 销售数据报告"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # 日期范围
        ws.merge_cells("A2:H2")
        ws["A2"] = f"日期范围: {start_date} ~ {end_date}"
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A2"].font = Font(color="666666", size=10)

        # 汇总统计
        ws["A4"] = "汇总统计"
        ws["A4"].font = Font(bold=True, size=12)

        summary_headers = ["指标", "数值"]
        summary_data = [
            ("点击数", merchant_report["clicks"] or 0 if merchant_report else 0),
            (
                "详情浏览",
                merchant_report["detail_views"] or 0 if merchant_report else 0,
            ),
            (
                "加购物车",
                merchant_report["add_to_carts"] or 0 if merchant_report else 0,
            ),
            ("购买数", merchant_report["purchases"] or 0 if merchant_report else 0),
            (
                "销售金额 ($)",
                float(merchant_report["amount"] or 0) if merchant_report else 0.0,
            ),
            (
                "佣金 ($)",
                float(merchant_report["commission"] or 0) if merchant_report else 0.0,
            ),
        ]

        for col, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, (label, value) in enumerate(summary_data, 6):
            ws.cell(row=row_idx, column=1, value=label).border = thin_border
            cell = ws.cell(row=row_idx, column=2, value=value)
            cell.border = thin_border
            cell.alignment = number_alignment
            if label in ["销售金额 ($)", "佣金 ($)"]:
                cell.fill = highlight_fill
                cell.number_format = "#,##0.00"
            else:
                cell.number_format = "#,##0"

        # 商品明细
        ws["A13"] = "商品明细"
        ws["A13"].font = Font(bold=True, size=12)

        product_headers = [
            "ASIN",
            "商品名称",
            "点击",
            "详情浏览",
            "加购",
            "购买",
            "销售金额 ($)",
            "佣金 ($)",
        ]
        for col, header in enumerate(product_headers, 1):
            cell = ws.cell(row=14, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, p in enumerate(product_reports, 15):
            ws.cell(row=row_idx, column=1, value=p["asin"] or "").border = thin_border
            ws.cell(
                row=row_idx, column=2, value=p["product_name"] or ""
            ).border = thin_border
            ws.cell(row=row_idx, column=3, value=p["clicks"] or 0).border = thin_border
            ws.cell(
                row=row_idx, column=4, value=p["detail_views"] or 0
            ).border = thin_border
            ws.cell(
                row=row_idx, column=5, value=p["add_to_carts"] or 0
            ).border = thin_border
            ws.cell(
                row=row_idx, column=6, value=p["purchases"] or 0
            ).border = thin_border
            ws.cell(
                row=row_idx, column=7, value=float(p["amount"] or 0)
            ).border = thin_border
            ws.cell(
                row=row_idx, column=8, value=float(p["commission"] or 0)
            ).border = thin_border

            # 数字格式
            for col in [3, 4, 5, 6]:
                ws.cell(row=row_idx, column=col).number_format = "#,##0"
                ws.cell(row=row_idx, column=col).alignment = number_alignment
            for col in [7, 8]:
                ws.cell(row=row_idx, column=col).number_format = "#,##0.00"
                ws.cell(row=row_idx, column=col).alignment = number_alignment
                ws.cell(row=row_idx, column=col).fill = highlight_fill

        # 调整列宽
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 40
        for col in ["C", "D", "E", "F"]:
            ws.column_dimensions[col].width = 12
        for col in ["G", "H"]:
            ws.column_dimensions[col].width = 15

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 生成文件名
        safe_name = re.sub(r'[\\/:*?"<>|]', "", merchant_name)
        filename = f"{safe_name}_销售数据_{start_date}_{end_date}.xlsx"

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# 全局状态
optimized_uploads = {}
_yp_sync_proc = None


@bp.route("/api/optimize/upload", methods=["POST"])
def api_optimize_upload():
    """
    T-002: 上传 Google Ads 报告文件（CSV或Excel）
    接收字段: file（文件）、asin（可选）、report_type
    """
    try:
        if "file" not in request.files:
            return jsonify({"ok": False, "msg": "未上传文件"})

        file = request.files["file"]
        asin = request.form.get("asin", "").strip()
        report_type = request.form.get("report_type", "search_term").strip()

        if not file or file.filename == "":
            return jsonify({"ok": False, "msg": "文件为空"})

        upload_id = str(uuid.uuid4())
        timestamp = datetime.datetime.now()

        rows = []

        # 判断文件类型并解析
        filename = file.filename.lower()
        if filename.endswith(".csv"):
            # CSV 解析
            try:
                content = file.read().decode("utf-8-sig")
                reader = csv.DictReader(io.StringIO(content))

                for row_dict in reader:
                    rows.append(row_dict)
            except UnicodeDecodeError:
                # 尝试 GBK 编码
                file.seek(0)
                content = file.read().decode("gbk")
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)

        elif filename.endswith((".xlsx", ".xls")):
            # Excel 解析
            try:
                import openpyxl

                file.seek(0)
                wb = openpyxl.load_workbook(file)
                ws = wb.active

                # 读取表头
                headers = [cell.value for cell in ws[1]]

                # 建立列索引映射
                col_mapping = {}
                for i, h in enumerate(headers):
                    std_col = _parse_report_column(str(h) if h else "")
                    if std_col:
                        col_mapping[i] = std_col

                # 读取数据行
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = {}
                    for idx, val in enumerate(row):
                        if idx in col_mapping:
                            row_dict[col_mapping[idx]] = val
                    if row_dict:
                        rows.append(row_dict)

            except ImportError:
                return jsonify(
                    {
                        "ok": False,
                        "msg": "Excel解析失败，请安装 openpyxl: pip install openpyxl",
                    }
                )
            except Exception as e:
                return jsonify({"ok": False, "msg": f"Excel解析错误: {e}"})

        else:
            return jsonify({"ok": False, "msg": "仅支持CSV或Excel文件"})

        if not rows:
            return jsonify({"ok": False, "msg": "文件中没有数据"})

        # 批量插入数据库
        conn = get_db()
        cur = conn.cursor()

        insert_sql = """
            INSERT INTO ads_search_term_reports 
            (upload_id, asin, report_type, search_term, impressions, clicks, ctr, cpc, cost,
             conversions, conv_rate, conv_value, quality_score, campaign_name, ad_group_name, match_type,
             upload_time)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        batch_data = []
        for row in rows:
            try:
                batch_data.append(
                    (
                        upload_id,
                        asin or None,
                        report_type,
                        row.get("search_term"),
                        _clean_numeric_value(row.get("impressions"), "int"),
                        _clean_numeric_value(row.get("clicks"), "int"),
                        _clean_numeric_value(row.get("ctr"), "float"),
                        _clean_numeric_value(row.get("cpc"), "float"),
                        _clean_numeric_value(row.get("cost"), "float"),
                        _clean_numeric_value(row.get("conversions"), "int"),
                        _clean_numeric_value(row.get("conv_rate"), "float"),
                        _clean_numeric_value(row.get("conv_value"), "float"),
                        _clean_numeric_value(row.get("quality_score"), "int"),
                        row.get("campaign_name"),
                        row.get("ad_group_name"),
                        row.get("match_type"),
                        timestamp,
                    )
                )
            except Exception:
                continue

        if batch_data:
            cur.executemany(insert_sql, batch_data)

        # 插入上传记录
        insert_upload_sql = """
            INSERT INTO ads_report_uploads 
            (upload_id, asin, report_type, file_name, row_count, status, upload_time)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        cur.execute(
            insert_upload_sql,
            (
                upload_id,
                asin or None,
                report_type,
                file.filename,
                len(batch_data),
                "pending",
                timestamp,
            ),
        )

        conn.commit()
        cur.close()
        conn.close()

        return jsonify(
            {
                "ok": True,
                "upload_id": upload_id,
                "row_count": len(batch_data),
                "msg": f"成功导入 {len(batch_data)} 条数据",
            }
        )

    except Exception as e:
        return jsonify({"ok": False, "msg": f"上传失败: {str(e)}"})


# ─── 广告优化模块（T-003: 分析引擎）───────────────────────────────────────


@bp.route("/api/optimize/analyze/<upload_id>", methods=["POST"])
def api_optimize_analyze(upload_id):
    """
    T-003: 分析上传的报告数据，生成优化建议
    """
    try:
        conn = get_db()
        cur = conn.cursor()

        # 查询本批次数据
        cur.execute(
            """
            SELECT id, search_term, impressions, clicks, ctr, cpc, cost,
                   conversions, conv_rate, conv_value, quality_score,
                   campaign_name, ad_group_name, match_type
            FROM ads_search_term_reports
            WHERE upload_id = %s
        """,
            (upload_id,),
        )

        rows = cur.fetchall()
        if not rows:
            cur.close()
            conn.close()
            return jsonify({"ok": False, "msg": "未找到该批次数据"})

        # 转换为字典列表
        data = []
        for row in rows:
            data.append(
                {
                    "id": row[0],
                    "search_term": row[1],
                    "impressions": row[2] or 0,
                    "clicks": row[3] or 0,
                    "ctr": row[4] or 0.0,
                    "cpc": row[5] or 0.0,
                    "cost": row[6] or 0.0,
                    "conversions": row[7] or 0,
                    "conv_rate": row[8] or 0.0,
                    "conv_value": row[9] or 0.0,
                    "quality_score": row[10] or 0,
                    "campaign_name": row[11],
                    "ad_group_name": row[12],
                    "match_type": row[13],
                }
            )

        suggestions = []
        total_impr = 0
        total_clicks = 0
        total_cost = 0.0
        total_conv = 0
        total_conv_value = 0.0

        # 分析每一条记录
        for item in data:
            total_impr += item["impressions"]
            total_clicks += item["clicks"]
            total_cost += item["cost"]
            total_conv += item["conversions"]
            total_conv_value += item["conv_value"]

            clicks = item["clicks"]
            conv_rate = item["conv_rate"]
            cost = item["cost"]
            conversions = item["conversions"]
            impressions = item["impressions"]
            quality_score = item["quality_score"]
            match_type = (item["match_type"] or "").lower()

            # 规则1：高点击低转化（否定词建议）
            if clicks >= 5 and conv_rate < 0.02 and cost > 2.0:
                priority = "high" if clicks >= 10 else "medium"
                suggestions.append(
                    {
                        "upload_id": upload_id,
                        "search_term": item["search_term"],
                        "rule_type": "add_negative",
                        "priority": priority,
                        "current_value": f"clicks={clicks}, conv_rate={conv_rate:.1%}, cost=${cost:.2f}",
                        "suggested_value": f"否定: {item['search_term']}",
                        "reason": f"该词 {clicks} 次点击，转化率仅 {conv_rate:.1%}，消耗 ${cost:.2f}，建议加入否定关键词",
                    }
                )

            # 规则2：高转化低曝光（提升出价建议）
            if conversions >= 1 and impressions < 100:
                suggestions.append(
                    {
                        "upload_id": upload_id,
                        "search_term": item["search_term"],
                        "rule_type": "increase_bid",
                        "priority": "high",
                        "current_value": f"conversions={conversions}, impressions={impressions}",
                        "suggested_value": "提升出价 20-30%",
                        "reason": f"该词已有 {conversions} 次转化但曝光仅 {impressions}，建议提升出价扩大曝光",
                    }
                )

            # 规则3：高转化可扩展匹配
            if conversions >= 2 and match_type in ("exact", "完全匹配"):
                suggestions.append(
                    {
                        "upload_id": upload_id,
                        "search_term": item["search_term"],
                        "rule_type": "expand_match",
                        "priority": "medium",
                        "current_value": f"conversions={conversions}, match_type={item['match_type']}",
                        "suggested_value": f'词组匹配: "{item["search_term"]}"',
                        "reason": f"高转化完全匹配词，建议同步添加词组匹配版本扩大覆盖",
                    }
                )

            # 规则4：高价值词建议独立Ad Group
            if conversions >= 3:
                suggestions.append(
                    {
                        "upload_id": upload_id,
                        "search_term": item["search_term"],
                        "rule_type": "new_ad_group",
                        "priority": "high",
                        "current_value": f"conversions={conversions}",
                        "suggested_value": f"新建Ad Group: [Exact] {item['search_term']}",
                        "reason": f"该词转化 {conversions} 次，建议单独成组精细化管理",
                    }
                )

            # 规则5：低质量分广告文案优化
            if quality_score > 0 and quality_score <= 5:
                suggestions.append(
                    {
                        "upload_id": upload_id,
                        "search_term": item["search_term"],
                        "rule_type": "rewrite_headline",
                        "priority": "high",
                        "current_value": f"quality_score={quality_score}/10",
                        "suggested_value": "优化广告标题和描述，提高相关性",
                        "reason": f"关键词质量得分仅 {quality_score}/10，建议优化广告文案提升相关性",
                    }
                )

            # 规则6：零转化高消耗词暂停建议
            if cost >= 10.0 and conversions == 0 and clicks >= 10:
                suggestions.append(
                    {
                        "upload_id": upload_id,
                        "search_term": item["search_term"],
                        "rule_type": "pause_keyword",
                        "priority": "high",
                        "current_value": f"cost=${cost:.2f}, clicks={clicks}, conversions=0",
                        "suggested_value": "暂停关键词",
                        "reason": f"已消耗 ${cost:.2f}，{clicks} 次点击零转化，建议暂停或加入否定",
                    }
                )

        # 插入优化建议
        if suggestions:
            insert_sugg_sql = """
                INSERT INTO ads_optimization_suggestions
                (upload_id, search_term, rule_type, priority, current_value, suggested_value, reason, created_at, status)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            now = datetime.datetime.now()
            batch_sugg = []
            for s in suggestions:
                batch_sugg.append(
                    (
                        s["upload_id"],
                        s["search_term"],
                        s["rule_type"],
                        s["priority"],
                        s["current_value"],
                        s["suggested_value"],
                        s["reason"],
                        now,
                        "pending",
                    )
                )
            cur.executemany(insert_sugg_sql, batch_sugg)

        # 计算KPI
        total_ctr = (total_clicks / total_impr * 100) if total_impr > 0 else 0.0
        total_roas = (total_conv_value / total_cost) if total_cost > 0 else 0.0

        # 写入KPI实际值
        kpi_metrics = [
            ("impressions", total_impr),
            ("clicks", total_clicks),
            ("ctr", round(total_ctr, 2)),
            ("cost", round(total_cost, 2)),
            ("conversions", total_conv),
            ("roas", round(total_roas, 2)),
            ("conv_value", round(total_conv_value, 2)),
        ]

        insert_kpi_sql = """
            INSERT INTO ads_kpi_actuals
            (upload_id, metric_name, actual_value, record_time)
            VALUES (%s, %s, %s, %s)
        """
        now = datetime.datetime.now()
        batch_kpi = []
        for metric_name, value in kpi_metrics:
            batch_kpi.append((upload_id, metric_name, value, now))
        cur.executemany(insert_kpi_sql, batch_kpi)

        # 更新上传记录状态
        cur.execute(
            """
            UPDATE ads_report_uploads
            SET status = 'done', suggestion_count = %s, analyze_time = %s
            WHERE upload_id = %s
        """,
            (len(suggestions), now, upload_id),
        )

        conn.commit()
        cur.close()
        conn.close()

        return jsonify(
            {
                "ok": True,
                "suggestions": len(suggestions),
                "kpi": {
                    "impressions": total_impr,
                    "clicks": total_clicks,
                    "ctr": round(total_ctr, 2),
                    "cost": round(total_cost, 2),
                    "conversions": total_conv,
                    "roas": round(total_roas, 2),
                    "conv_value": round(total_conv_value, 2),
                },
                "msg": f"分析完成，生成 {len(suggestions)} 条优化建议",
            }
        )

    except Exception as e:
        return jsonify({"ok": False, "msg": f"分析失败: {str(e)}"})


@bp.route("/api/optimize/upload_excel", methods=["POST"])
def api_optimize_upload_excel():
    """上传Excel报告文件并解析"""
    try:
        if "file" not in request.files:
            return jsonify({"success": False, "error": "未上传文件"})

        file = request.files["file"]
        report_type = request.form.get("report_type", "keywords")
        merchant_id = request.form.get("merchant_id", "")
        start_date = request.form.get("start_date", "")
        end_date = request.form.get("end_date", "")

        if not file or file.filename == "":
            return jsonify({"success": False, "error": "文件为空"})

        if not merchant_id:
            return jsonify({"success": False, "error": "请选择品牌"})

        # 解析Excel
        rows = []
        filename = file.filename.lower()

        if filename.endswith((".xlsx", ".xls")):
            try:
                import openpyxl

                file.seek(0)
                wb = openpyxl.load_workbook(file)
                ws = wb.active

                # 读取表头
                headers = [
                    str(cell.value).strip() if cell.value else "" for cell in ws[1]
                ]

                # 读取数据行
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = {}
                    for idx, val in enumerate(row):
                        if idx < len(headers) and headers[idx]:
                            row_dict[headers[idx]] = val
                    if row_dict:
                        rows.append(row_dict)

            except Exception as e:
                return jsonify({"success": False, "error": f"Excel解析错误: {e}"})
        else:
            return jsonify({"success": False, "error": "仅支持Excel文件"})

        if not rows:
            return jsonify({"success": False, "error": "文件中没有数据"})

        # 根据报告类型处理数据
        data = []
        for row in rows:
            if report_type == "keywords":
                # 搜索关键字报告
                data.append(
                    {
                        "keyword": row.get("Keyword")
                        or row.get("关键字")
                        or row.get("Search keyword")
                        or "",
                        "match_type": row.get("Match type")
                        or row.get("匹配类型")
                        or "",
                        "campaign": row.get("Campaign") or row.get("广告系列") or "",
                        "ad_group": row.get("Ad group") or row.get("广告组") or "",
                        "impressions": int(
                            float(row.get("Impr.") or row.get("展示次数") or 0)
                        ),
                        "clicks": int(
                            float(row.get("Clicks") or row.get("点击次数") or 0)
                        ),
                        "ctr": float(
                            str(row.get("CTR") or row.get("点击率") or "0").replace(
                                "%", ""
                            )
                        )
                        / 100,
                        "cpc": float(row.get("Avg. CPC") or row.get("平均CPC") or 0),
                        "cost": float(row.get("Cost") or row.get("花费") or 0),
                        "conversions": int(
                            float(row.get("Conversions") or row.get("转化") or 0)
                        ),
                    }
                )
            elif report_type == "search_terms":
                # 搜索字词报告
                data.append(
                    {
                        "search_term": row.get("Search term")
                        or row.get("搜索词")
                        or row.get("Search query")
                        or "",
                        "keyword": row.get("Keyword") or row.get("关键字") or "",
                        "match_type": row.get("Match type")
                        or row.get("匹配类型")
                        or "",
                        "campaign": row.get("Campaign") or row.get("广告系列") or "",
                        "ad_group": row.get("Ad group") or row.get("广告组") or "",
                        "impressions": int(
                            float(row.get("Impr.") or row.get("展示次数") or 0)
                        ),
                        "clicks": int(
                            float(row.get("Clicks") or row.get("点击次数") or 0)
                        ),
                        "ctr": float(
                            str(row.get("CTR") or row.get("点击率") or "0").replace(
                                "%", ""
                            )
                        )
                        / 100,
                        "cpc": float(row.get("Avg. CPC") or row.get("平均CPC") or 0),
                        "cost": float(row.get("Cost") or row.get("花费") or 0),
                        "conversions": int(
                            float(row.get("Conversions") or row.get("转化") or 0)
                        ),
                    }
                )

        # 调用投放数据API保存
        from routes_products import api_workflow_ads_data

        # 构造请求对象
        class FakeRequest:
            def get_json(self, silent=False):
                return {
                    "report_type": report_type,
                    "report_date": end_date,
                    "start_date": start_date,
                    "end_date": end_date,
                    "data": data,
                }

        # 直接保存到数据库
        conn = get_db()
        cur = conn.cursor()

        saved = 0
        if report_type == "keywords":
            for row in data:
                if not row.get("keyword"):
                    continue
                try:
                    cur.execute(
                        """
                        INSERT INTO ads_keywords_report
                        (merchant_id, report_date, keyword, match_type, campaign, ad_group,
                         impressions, clicks, ctr, cpc, cost, conversions)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                        (
                            merchant_id,
                            end_date,
                            row.get("keyword", ""),
                            row.get("match_type", ""),
                            row.get("campaign", ""),
                            row.get("ad_group", ""),
                            row.get("impressions", 0),
                            row.get("clicks", 0),
                            row.get("ctr", 0),
                            row.get("cpc", 0),
                            row.get("cost", 0),
                            row.get("conversions", 0),
                        ),
                    )
                    saved += 1
                except Exception as e:
                    print(f"[ERROR] 保存关键字失败: {e}")
        elif report_type == "search_terms":
            for row in data:
                if not row.get("search_term"):
                    continue
                try:
                    cur.execute(
                        """
                        INSERT INTO ads_search_terms_report
                        (merchant_id, report_date, search_term, keyword, match_type, campaign, ad_group,
                         impressions, clicks, ctr, cpc, cost, conversions)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                        (
                            merchant_id,
                            end_date,
                            row.get("search_term", ""),
                            row.get("keyword", ""),
                            row.get("match_type", ""),
                            row.get("campaign", ""),
                            row.get("ad_group", ""),
                            row.get("impressions", 0),
                            row.get("clicks", 0),
                            row.get("ctr", 0),
                            row.get("cpc", 0),
                            row.get("cost", 0),
                            row.get("conversions", 0),
                        ),
                    )
                    saved += 1
                except Exception as e:
                    print(f"[ERROR] 保存搜索词失败: {e}")

        conn.commit()
        cur.close()
        conn.close()

        return jsonify(
            {"success": True, "saved_count": saved, "message": f"已保存 {saved} 条数据"}
        )

    except Exception as e:
        import traceback

        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


# ─── 广告优化模块（T-004: 数据查询API）────────────────────────────────────


@bp.route("/api/optimize/uploads", methods=["GET"])
def api_optimize_uploads():
    """返回所有上传记录列表"""
    try:
        conn = get_db()
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT upload_id, asin, report_type, file_name, row_count, 
                   status, suggestion_count, upload_time, analyze_time
            FROM ads_report_uploads
            ORDER BY upload_time DESC
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        # 转换 datetime 为字符串
        for row in rows:
            for key, val in row.items():
                if isinstance(val, datetime.datetime):
                    row[key] = val.strftime("%Y-%m-%d %H:%M:%S")

        return jsonify({"ok": True, "data": rows})

    except Exception as e:
        return jsonify({"ok": False, "msg": f"查询失败: {str(e)}"})


@bp.route("/api/optimize/suggestions/<upload_id>", methods=["GET"])
def api_optimize_suggestions(upload_id):
    """返回该批次所有优化建议（按priority排序）"""
    try:
        conn = get_db()
        cur = conn.cursor(dictionary=True)
        cur.execute(
            """
            SELECT id, search_term, rule_type, priority, current_value, 
                   suggested_value, reason, status, created_at
            FROM ads_optimization_suggestions
            WHERE upload_id = %s
            ORDER BY FIELD(priority, 'high', 'medium', 'low') DESC, created_at DESC
        """,
            (upload_id,),
        )
        rows = cur.fetchall()
        cur.close()
        conn.close()

        for row in rows:
            for key, val in row.items():
                if isinstance(val, datetime.datetime):
                    row[key] = val.strftime("%Y-%m-%d %H:%M:%S")

        return jsonify({"ok": True, "data": rows})

    except Exception as e:
        return jsonify({"ok": False, "msg": f"查询失败: {str(e)}"})


@bp.route("/api/optimize/kpi", methods=["GET"])
def api_optimize_kpi():
    """返回所有KPI目标 + 最新实际值对比"""
    try:
        conn = get_db()
        cur = conn.cursor(dictionary=True)

        # 获取KPI目标
        cur.execute("""
            SELECT id, metric_name, target_value, alert_threshold, asin, campaign_name, created_at
            FROM ads_kpi_targets
            ORDER BY created_at DESC
        """)
        targets = cur.fetchall()

        # 获取最新实际值
        cur.execute("""
            SELECT t1.upload_id, t1.metric_name, t1.actual_value, t1.record_time
            FROM ads_kpi_actuals t1
            INNER JOIN (
                SELECT metric_name, MAX(record_time) as max_time
                FROM ads_kpi_actuals
                GROUP BY metric_name
            ) t2 ON t1.metric_name = t2.metric_name AND t1.record_time = t2.max_time
        """)
        actuals = cur.fetchall()

        cur.close()
        conn.close()

        # 构建实际值映射
        actual_map = {}
        for act in actuals:
            actual_map[act["metric_name"]] = act["actual_value"]

        # 合并数据
        result = []
        for tgt in targets:
            metric = tgt["metric_name"]
            actual_val = actual_map.get(metric, None)

            result.append(
                {
                    "target_id": tgt["id"],
                    "metric_name": metric,
                    "target_value": tgt["target_value"],
                    "actual_value": actual_val,
                    "alert_threshold": tgt["alert_threshold"],
                    "asin": tgt["asin"],
                    "campaign_name": tgt["campaign_name"],
                    "is_alert": actual_val is not None
                    and abs(actual_val - tgt["target_value"]) > tgt["alert_threshold"],
                }
            )

        return jsonify({"ok": True, "data": result})

    except Exception as e:
        return jsonify({"ok": False, "msg": f"查询失败: {str(e)}"})


@bp.route("/api/optimize/kpi_target", methods=["POST"])
def api_optimize_kpi_target():
    """保存KPI目标"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"ok": False, "msg": "缺少数据"})

        metric_name = data.get("metric_name", "").strip()
        target_value = data.get("target_value")
        alert_threshold = data.get("alert_threshold", 0)
        asin = data.get("asin", "").strip() or None
        campaign_name = data.get("campaign_name", "").strip() or None

        if not metric_name or target_value is None:
            return jsonify({"ok": False, "msg": "metric_name 和 target_value 为必填项"})

        conn = get_db()
        cur = conn.cursor()

        insert_sql = """
            INSERT INTO ads_kpi_targets
            (metric_name, target_value, alert_threshold, asin, campaign_name, created_at)
            VALUES (%s, %s, %s, %s, %s, %s)
        """
        cur.execute(
            insert_sql,
            (
                metric_name,
                float(target_value),
                float(alert_threshold),
                asin,
                campaign_name,
                datetime.datetime.now(),
            ),
        )

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({"ok": True, "msg": "KPI目标已保存"})

    except Exception as e:
        return jsonify({"ok": False, "msg": f"保存失败: {str(e)}"})


@bp.route("/api/optimize/suggestion_action", methods=["POST"])
def api_optimize_suggestion_action():
    """更新优化建议状态（应用/忽略）"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"ok": False, "msg": "缺少数据"})

        suggestion_id = data.get("id")
        status = data.get("status", "").strip()

        if not suggestion_id or not status:
            return jsonify({"ok": False, "msg": "id 和 status 为必填项"})

        if status not in ("applied", "dismissed"):
            return jsonify({"ok": False, "msg": "status 必须为 applied 或 dismissed"})

        conn = get_db()
        cur = conn.cursor()

        update_sql = """
            UPDATE ads_optimization_suggestions
            SET status = %s, updated_at = %s
            WHERE id = %s
        """
        cur.execute(update_sql, (status, datetime.datetime.now(), suggestion_id))

        conn.commit()
        affected = cur.rowcount
        cur.close()
        conn.close()

        if affected == 0:
            return jsonify({"ok": False, "msg": "未找到该建议"})

        return jsonify({"ok": True, "msg": f"建议状态已更新为 {status}"})

    except Exception as e:
        return jsonify({"ok": False, "msg": f"更新失败: {str(e)}"})


# ═══════════════════════════════════════════════════════════════════════════
# T-007~T-010: 广告质量评分（Quality Score）后端
# ═══════════════════════════════════════════════════════════════════════════


def _ensure_quality_score_column():
    """T-007: 确保 ads_ads 和 ads_plans 表中存在 quality_score 字段"""
    try:
        conn = get_db()
        cur = conn.cursor()
        # ads_ads 表加 quality_score
        cur.execute("""
            SELECT COUNT(*) FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE()
              AND TABLE_NAME = 'ads_ads'
              AND COLUMN_NAME = 'quality_score'
        """)
        if cur.fetchone()[0] == 0:
            cur.execute(
                "ALTER TABLE ads_ads ADD COLUMN quality_score FLOAT DEFAULT NULL COMMENT 'QS评分0-100'"
            )
            cur.execute(
                "ALTER TABLE ads_ads ADD COLUMN qs_detail JSON DEFAULT NULL COMMENT 'QS各维度明细'"
            )
            conn.commit()
        # ads_plans 表加 avg_quality_score
        cur.execute("""
            SELECT COUNT(*) FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE()
              AND TABLE_NAME = 'ads_plans'
              AND COLUMN_NAME = 'avg_quality_score'
        """)
        if cur.fetchone()[0] == 0:
            cur.execute(
                "ALTER TABLE ads_plans ADD COLUMN avg_quality_score FLOAT DEFAULT NULL COMMENT '方案平均QS'"
            )
            conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        app.logger.warning(f"_ensure_quality_score_column: {e}")


def calculate_quality_score(ad_data):
    """T-008: 计算单条广告的 Quality Score（0-100）

    评分维度（共100分）：
    - 标题质量     (30分): 数量/长度/含关键词/含数字/含CTA
    - 描述质量     (25分): 数量/长度/含USP/字符合规
    - 扩展资产     (20分): sitelinks/callouts/structured_snippet
    - 合规性       (15分): 所有字段字符数是否合规
    - 文案多样性   (10分): A/B变体区别度
    """
    import json as _json

    score = 0.0
    detail = {}

    # 解析 JSON 字段
    try:
        headlines = (
            _json.loads(ad_data.get("headlines", "[]"))
            if isinstance(ad_data.get("headlines"), str)
            else (ad_data.get("headlines") or [])
        )
    except Exception:
        headlines = []
    try:
        descriptions = (
            _json.loads(ad_data.get("descriptions", "[]"))
            if isinstance(ad_data.get("descriptions"), str)
            else (ad_data.get("descriptions") or [])
        )
    except Exception:
        descriptions = []
    try:
        sitelinks = (
            _json.loads(ad_data.get("sitelinks", "[]"))
            if isinstance(ad_data.get("sitelinks"), str)
            else (ad_data.get("sitelinks") or [])
        )
    except Exception:
        sitelinks = []
    try:
        callouts = (
            _json.loads(ad_data.get("callouts", "[]"))
            if isinstance(ad_data.get("callouts"), str)
            else (ad_data.get("callouts") or [])
        )
    except Exception:
        callouts = []
    try:
        snippets = (
            _json.loads(ad_data.get("structured_snippet", "{}"))
            if isinstance(ad_data.get("structured_snippet"), str)
            else (ad_data.get("structured_snippet") or {})
        )
    except Exception:
        snippets = {}

    # ── 1. 标题质量 (30分) ──
    hl_texts = []
    for h in headlines:
        if isinstance(h, dict):
            hl_texts.append(h.get("text", ""))
        elif isinstance(h, str):
            hl_texts.append(h)
    hl_count = len(hl_texts)
    hl_score = 0.0
    # 数量分：3条=6，5条=10，8条=14，≥10条=16
    hl_count_map = {0: 0, 1: 2, 2: 4, 3: 6, 4: 8, 5: 10, 6: 11, 7: 12, 8: 14, 9: 15}
    hl_score += hl_count_map.get(min(hl_count, 9), 16) if hl_count < 10 else 16
    # 长度分：平均长度 ≥ 20 加4分
    if hl_texts:
        avg_len = sum(len(t) for t in hl_texts) / len(hl_texts)
        hl_score += min(avg_len / 30.0 * 8.0, 8.0)
    # 合规分：所有标题 ≤ 30 字符
    hl_valid = all(len(t) <= 30 for t in hl_texts) if hl_texts else False
    hl_score += 6.0 if hl_valid else 0.0
    score += min(hl_score, 30.0)
    detail["headline"] = round(min(hl_score, 30.0), 1)

    # ── 2. 描述质量 (25分) ──
    desc_texts = []
    for d in descriptions:
        if isinstance(d, dict):
            desc_texts.append(d.get("text", ""))
        elif isinstance(d, str):
            desc_texts.append(d)
    desc_count = len(desc_texts)
    desc_score = 0.0
    # 数量分
    desc_count_map = {0: 0, 1: 5, 2: 10, 3: 12, 4: 14}
    desc_score += desc_count_map.get(min(desc_count, 4), 14)
    # 长度分：平均 ≥ 70 字符
    if desc_texts:
        avg_dlen = sum(len(t) for t in desc_texts) / len(desc_texts)
        desc_score += min(avg_dlen / 90.0 * 7.0, 7.0)
    # 合规分：所有描述 ≤ 90 字符
    desc_valid = all(len(t) <= 90 for t in desc_texts) if desc_texts else False
    desc_score += 4.0 if desc_valid else 0.0
    score += min(desc_score, 25.0)
    detail["description"] = round(min(desc_score, 25.0), 1)

    # ── 3. 扩展资产 (20分) ──
    asset_score = 0.0
    # Sitelinks：有2条+6分，有4条+10分
    sl_count = len(sitelinks) if isinstance(sitelinks, list) else 0
    asset_score += min(sl_count / 4.0 * 10.0, 10.0)
    # Callouts：有2条+4分，有4条+6分
    co_count = len(callouts) if isinstance(callouts, list) else 0
    asset_score += min(co_count / 4.0 * 6.0, 6.0)
    # Structured Snippet
    if (
        snippets
        and (isinstance(snippets, dict) and snippets.get("values"))
        or isinstance(snippets, list)
    ):
        asset_score += 4.0
    score += min(asset_score, 20.0)
    detail["assets"] = round(min(asset_score, 20.0), 1)

    # ── 4. 合规性 (15分) ──
    compliance_score = 15.0
    # 有违规标题扣分
    bad_hl = [t for t in hl_texts if len(t) > 30]
    bad_desc = [t for t in desc_texts if len(t) > 90]
    compliance_score -= len(bad_hl) * 3.0
    compliance_score -= len(bad_desc) * 4.0
    compliance_score = max(compliance_score, 0.0)
    # all_chars_valid 字段
    if ad_data.get("all_chars_valid"):
        compliance_score = min(compliance_score + 3.0, 15.0)
    score += compliance_score
    detail["compliance"] = round(compliance_score, 1)

    # ── 5. 文案多样性 (10分) ──
    diversity_score = 5.0  # 基础分
    if hl_count >= 10 and desc_count >= 4:
        diversity_score = 10.0
    elif hl_count >= 8 and desc_count >= 3:
        diversity_score = 8.0
    elif hl_count >= 5 and desc_count >= 2:
        diversity_score = 7.0
    score += diversity_score
    detail["diversity"] = round(diversity_score, 1)

    total = round(min(score, 100.0), 1)
    detail["total"] = total
    return total, detail


@bp.route("/api/ads/score/<asin>", methods=["POST"])
def api_ads_score_asin(asin):
    """T-009: 对指定 ASIN 的所有广告计算 QS 评分，更新 ads_ads 表，返回平均分"""
    try:
        _ensure_quality_score_column()
        conn = get_db()
        cur = conn.cursor(dictionary=True)
        cur.execute(
            """
            SELECT a.id, a.variant, a.headlines, a.descriptions,
                   a.sitelinks, a.callouts, a.structured_snippet,
                   a.all_chars_valid
            FROM ads_ads a
            WHERE a.asin = %s
        """,
            (asin,),
        )
        ads = cur.fetchall()
        if not ads:
            cur.close()
            conn.close()
            return jsonify({"ok": False, "msg": f"未找到 ASIN={asin} 的广告"})

        results = []
        import json as _json

        for ad in ads:
            qs, detail = calculate_quality_score(ad)
            cur.execute(
                """
                UPDATE ads_ads SET quality_score=%s, qs_detail=%s, updated_at=NOW()
                WHERE id=%s
            """,
                (qs, _json.dumps(detail), ad["id"]),
            )
            results.append(
                {
                    "ad_id": ad["id"],
                    "variant": ad["variant"],
                    "qs": qs,
                    "detail": detail,
                }
            )

        # 更新 ads_plans 的平均分
        avg_qs = sum(r["qs"] for r in results) / len(results) if results else 0
        cur.execute(
            """
            UPDATE ads_plans SET avg_quality_score=%s, updated_at=NOW()
            WHERE asin=%s
        """,
            (round(avg_qs, 1), asin),
        )

        conn.commit()
        cur.close()
        conn.close()
        return jsonify(
            {"ok": True, "asin": asin, "avg_qs": round(avg_qs, 1), "ads": results}
        )

    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})


@bp.route("/api/ads/scores", methods=["GET"])
def api_ads_scores_all():
    """T-009: 获取所有 ASIN 的平均 QS 评分列表（用于 /qs_dashboard）"""
    try:
        _ensure_quality_score_column()
        conn = get_db()
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT p.asin, p.product_name, p.avg_quality_score,
                   p.ad_count,
                   COUNT(a.id) AS scored_ads,
                   SUM(CASE WHEN a.quality_score >= 80 THEN 1 ELSE 0 END) AS qs_good,
                   SUM(CASE WHEN a.quality_score >= 60 AND a.quality_score < 80 THEN 1 ELSE 0 END) AS qs_ok,
                   SUM(CASE WHEN a.quality_score < 60 THEN 1 ELSE 0 END) AS qs_bad
            FROM ads_plans p
            LEFT JOIN ads_ads a ON a.asin = p.asin
            WHERE p.plan_status = 'completed'
            GROUP BY p.asin, p.product_name, p.avg_quality_score, p.ad_count
            ORDER BY p.avg_quality_score ASC
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        # 汇总统计
        total_ads = sum(r["scored_ads"] or 0 for r in rows)
        qs_good_total = sum(r["qs_good"] or 0 for r in rows)
        qs_bad_total = sum(r["qs_bad"] or 0 for r in rows)
        avg_overall = (
            sum((r["avg_quality_score"] or 0) for r in rows) / len(rows) if rows else 0
        )

        return jsonify(
            {
                "ok": True,
                "summary": {
                    "total_asins": len(rows),
                    "total_ads": total_ads,
                    "avg_qs": round(avg_overall, 1),
                    "qs_good": qs_good_total,
                    "qs_bad": qs_bad_total,
                },
                "items": rows,
            }
        )
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})


@bp.route("/api/ads/score_batch", methods=["POST"])
def api_ads_score_batch():
    """T-009: 对所有 completed 方案批量计算 QS，后台更新"""
    try:
        _ensure_quality_score_column()
        conn = get_db()
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT DISTINCT asin FROM ads_plans WHERE plan_status='completed'")
        asins = [r["asin"] for r in cur.fetchall()]
        cur.close()
        conn.close()

        updated = 0
        import json as _json

        for asin in asins:
            conn2 = get_db()
            cur2 = conn2.cursor(dictionary=True)
            cur2.execute(
                """
                SELECT id, variant, headlines, descriptions,
                       sitelinks, callouts, structured_snippet, all_chars_valid
                FROM ads_ads WHERE asin=%s
            """,
                (asin,),
            )
            ads = cur2.fetchall()
            if not ads:
                cur2.close()
                conn2.close()
                continue
            scores = []
            for ad in ads:
                qs, detail = calculate_quality_score(ad)
                cur2.execute(
                    "UPDATE ads_ads SET quality_score=%s, qs_detail=%s, updated_at=NOW() WHERE id=%s",
                    (qs, _json.dumps(detail), ad["id"]),
                )
                scores.append(qs)
            avg_qs = sum(scores) / len(scores) if scores else 0
            cur2.execute(
                "UPDATE ads_plans SET avg_quality_score=%s, updated_at=NOW() WHERE asin=%s",
                (round(avg_qs, 1), asin),
            )
            conn2.commit()
            cur2.close()
            conn2.close()
            updated += 1

        return jsonify({"ok": True, "msg": f"已批量评分 {updated} 个 ASIN"})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})


# T-010: 在 Flask 启动时自动初始化 quality_score 字段（调用一次）
try:
    with app.app_context():
        _ensure_quality_score_column()
except Exception:
    pass


# ═══════════════════════════════════════════════════════════════════════════
# 投放优化模块 - Agent对话模式前端页面
# ═══════════════════════════════════════════════════════════════════════════

OPTIMIZE_HTML = (
    """
<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>投放优化 - YP Affiliate</title>
<style>
"""
    + BASE_CSS
    + """
/* 页面布局 */
.optimize-page { padding: 20px; max-width: 1400px; margin: 0 auto; }

/* 商户选择板块 */
.merchant-select-card {
  background: #1a1d24; border: 1px solid #2a2d36; border-radius: 12px;
  padding: 20px; margin-bottom: 20px;
}
.merchant-select-card h3 { font-size: 14px; color: #fff; margin-bottom: 16px; display: flex; align-items: center; gap: 8px; }
.merchant-row { display: grid; grid-template-columns: 1fr 1fr 1fr 1fr auto; gap: 12px; align-items: end; }
.merchant-row > div { display: flex; flex-direction: column; gap: 4px; }
.merchant-row label { font-size: 12px; color: #888; }
.merchant-row select, .merchant-row input {
  padding: 10px 12px; border: 1px solid #2a2d36; border-radius: 8px;
  background: #15181f; color: #fff; font-size: 14px;
}
.merchant-row select:focus, .merchant-row input:focus { outline: none; border-color: #ffa726; }
.btn-row { display: flex; gap: 8px; }
.btn-row button { white-space: nowrap; }

/* 销售数据板块 */
.sales-data-card {
  background: #1a1d24; border: 1px solid #2a2d36; border-radius: 12px;
  padding: 20px; margin-bottom: 20px;
}
.sales-data-card h3 { font-size: 14px; color: #fff; margin-bottom: 16px; display: flex; align-items: center; gap: 8px; }

/* 汇总统计 */
.summary-stats { display: grid; grid-template-columns: repeat(6, 1fr); gap: 12px; margin-bottom: 20px; }
.stat-item {
  background: #15181f; border-radius: 8px; padding: 16px; text-align: center;
}
.stat-item .stat-value { font-size: 24px; font-weight: 700; color: #fff; }
.stat-item .stat-label { font-size: 12px; color: #888; margin-top: 4px; }
.stat-item.highlight .stat-value { color: #4caf50; }
.stat-item.warning .stat-value { color: #ffa726; }

/* 商品表格 */
.products-table { width: 100%; border-collapse: collapse; font-size: 13px; }
.products-table th { background: #15181f; padding: 12px; text-align: left; border-bottom: 1px solid #2a2d36; color: #888; font-weight: 600; }
.products-table td { padding: 12px; border-bottom: 1px solid #2a2d36; }
.products-table tr:hover { background: #1e222a; }
.products-table .num { text-align: right; font-family: monospace; }
.products-table .asin { color: #64b5f6; }

/* ROI 板块 */
.roi-card {
  background: #1a1d24; border: 1px solid #2a2d36; border-radius: 12px;
  padding: 20px; margin-bottom: 20px;
}
.roi-card h3 { font-size: 14px; color: #fff; margin-bottom: 16px; display: flex; align-items: center; gap: 8px; }
.roi-content { display: grid; grid-template-columns: 250px 1fr; gap: 24px; align-items: center; }
.roi-input-section { display: flex; flex-direction: column; gap: 8px; }
.roi-input-section label { font-size: 12px; color: #888; }
.roi-input-section input {
  padding: 12px 16px; border: 2px solid #ffa726; border-radius: 8px;
  background: #15181f; color: #ffa726; font-size: 24px; font-weight: 700;
  text-align: center; width: 200px;
}
.roi-input-section input:focus { outline: none; border-color: #ff9800; }
.roi-hint { font-size: 11px; color: #666; }
.roi-result-section { display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; }
.roi-item {
  background: #15181f; border-radius: 8px; padding: 16px; text-align: center;
}
.roi-item .roi-label { font-size: 12px; color: #888; margin-bottom: 8px; }
.roi-item .roi-value { font-size: 20px; font-weight: 700; color: #fff; }
.roi-item.highlight { background: #1a2e1a; border: 1px solid #4caf50; }
.roi-item.highlight .roi-value { color: #4caf50; }
.roi-item.negative .roi-value { color: #ef5350; }

/* AI分析板块 */
.ai-analysis-card {
  background: #1a1d24; border: 1px solid #2a2d36; border-radius: 12px;
  padding: 20px;
}
.ai-analysis-card h3 { font-size: 14px; color: #fff; margin-bottom: 16px; display: flex; align-items: center; gap: 8px; }

/* 上传区域 */
.upload-row { display: grid; grid-template-columns: 1fr 1fr auto; gap: 12px; align-items: end; margin-bottom: 16px; }
.upload-item {
  background: #15181f; border: 1px dashed #2a2d36; border-radius: 8px;
  padding: 12px; cursor: pointer; transition: all .15s;
}
.upload-item:hover { border-color: #ffa726; }
.upload-item.has-file { border-color: #4caf50; border-style: solid; }
.upload-item input { display: none; }
.upload-item-text { font-size: 13px; color: #e0e0e0; }
.upload-item-status { font-size: 11px; color: #4caf50; margin-top: 4px; }

/* 花费输入 */
.cost-input { display: flex; flex-direction: column; gap: 4px; }
.cost-input label { font-size: 12px; color: #888; }
.cost-input input {
  padding: 10px 12px; border: 1px solid #2a2d36; border-radius: 8px;
  background: #15181f; color: #fff; font-size: 14px;
}

/* 分析结果 */
.analysis-result {
  background: #15181f; border-radius: 8px; padding: 16px;
  margin-top: 16px; white-space: pre-wrap; font-size: 13px; line-height: 1.6;
  max-height: 400px; overflow-y: auto;
}

/* 按钮 */
.btn { display: inline-flex; align-items: center; gap: 6px; padding: 10px 16px; border-radius: 8px; border: none; font-size: 13px; font-weight: 600; cursor: pointer; transition: all .15s; }
.btn-primary { background: #1565c0; color: #fff; }
.btn-primary:hover { background: #1976d2; }
.btn-success { background: #2e7d32; color: #fff; }
.btn-success:hover { background: #388e3c; }
.btn-secondary { background: #23262f; color: #adb5bd; border: 1px solid #2a2d36; }
.btn-secondary:hover { background: #2a2d36; color: #e0e0e0; }
.btn-sm { padding: 6px 12px; font-size: 12px; }
.btn:disabled { opacity: 0.5; cursor: not-allowed; }

/* 加载动画 */
.loading { text-align: center; padding: 40px; color: #888; }
.loading-spinner { width: 40px; height: 40px; border: 3px solid #2a2d36; border-top-color: #ffa726; border-radius: 50%; animation: spin 1s linear infinite; margin: 0 auto 16px; }
@keyframes spin { to { transform: rotate(360deg); } }

/* 空状态 */
.empty-state { text-align: center; padding: 60px 20px; color: #666; }
.empty-state-icon { font-size: 48px; margin-bottom: 16px; }

/* Google Ads 总花费板块 */
.total-cost-card {
  background: #1a1d24; border: 1px solid #2a2d36; border-radius: 12px;
  padding: 20px; margin-bottom: 20px;
}
.total-cost-card h3 { font-size: 14px; color: #fff; margin-bottom: 16px; display: flex; align-items: center; gap: 8px; }
.total-cost-stats { display: grid; grid-template-columns: repeat(3, 1fr); gap: 16px; }
.total-cost-item {
  background: #15181f; border-radius: 8px; padding: 16px; text-align: center;
}
.total-cost-item .total-cost-label { font-size: 12px; color: #888; margin-bottom: 8px; }
.total-cost-item .total-cost-value { font-size: 24px; font-weight: 700; color: #fff; }
.total-cost-item.highlight { background: #1a2e1a; border: 1px solid #4caf50; }
.total-cost-item.highlight .total-cost-value { color: #4caf50; }
.total-cost-item.negative .total-cost-value { color: #ef5350; }
</style>
</head>
<body>
"""
    + NAV_HTML.format(
        p0="",
        p1="",
        p2="",
        p3="",
        p4="",
        p5="",
        p6="active",
        p7="",
        p8="",
        p9="",
        p10="",
        p11="",
    )
    + """
<div class="optimize-page">
  <!-- Google Ads 总体数据板块 -->
  <div class="total-cost-card">
    <h3>📊 Google Ads 总体数据（所有品牌）</h3>
    <div class="total-cost-stats">
      <div class="total-cost-item">
        <div class="total-cost-label">总花费 (￥)</div>
        <div class="total-cost-value" id="total-ads-cost">￥0</div>
      </div>
      <div class="total-cost-item">
        <div class="total-cost-label">总佣金 (￥)</div>
        <div class="total-cost-value" id="total-commission">￥0</div>
      </div>
      <div class="total-cost-item highlight">
        <div class="total-cost-label">总利润 (￥)</div>
        <div class="total-cost-value" id="total-profit">￥0</div>
      </div>
    </div>
  </div>

  <!-- 商户选择板块 -->
  <div class="merchant-select-card">
    <h3>🏪 选择商户</h3>
    <div class="merchant-row">
      <div>
        <label>商户列表</label>
        <select id="merchant-select" onchange="onMerchantSelect()">
          <option value="">-- 选择商户 --</option>
        </select>
      </div>
      <div>
        <label>或输入商户ID</label>
        <input type="text" id="merchant-id-input" placeholder="如: 364290" onchange="onMerchantInput()">
      </div>
      <div>
        <label>开始日期</label>
        <input type="date" id="date-start">
      </div>
      <div>
        <label>结束日期</label>
        <input type="date" id="date-end">
      </div>
      <div class="btn-row">
        <button class="btn btn-primary" onclick="fetchSalesData()" id="fetch-btn" disabled>🔄 拉取数据</button>
        <button class="btn btn-secondary btn-sm" onclick="setDateRange('this_year')">今年</button>
        <button class="btn btn-secondary btn-sm" onclick="setDateRange('this_month')">本月</button>
        <button class="btn btn-secondary btn-sm" onclick="setDateRange('last_month')">上月</button>
        <button class="btn btn-secondary btn-sm" onclick="setDateRange('last_7days')">近7天</button>
      </div>
    </div>
    <div id="fetch-status" style="font-size:12px;color:#888;margin-top:12px;"></div>
  </div>

  <!-- 销售数据板块 -->
  <div class="sales-data-card" id="sales-card" style="display:none;">
    <h3>📊 销售数据 <span id="sales-date-range" style="color:#666;font-weight:normal;"></span>
      <button class="btn btn-success btn-sm" onclick="downloadExcel()" id="download-btn" style="margin-left:12px;">📥 下载 Excel</button>
    </h3>
    
    <!-- 汇总统计 -->
    <div class="summary-stats" id="summary-stats">
      <div class="stat-item">
        <div class="stat-value" id="stat-clicks">0</div>
        <div class="stat-label">点击数</div>
      </div>
      <div class="stat-item">
        <div class="stat-value" id="stat-views">0</div>
        <div class="stat-label">详情浏览</div>
      </div>
      <div class="stat-item">
        <div class="stat-value" id="stat-carts">0</div>
        <div class="stat-label">加购物车</div>
      </div>
      <div class="stat-item">
        <div class="stat-value" id="stat-purchases">0</div>
        <div class="stat-label">购买数</div>
      </div>
      <div class="stat-item highlight">
        <div class="stat-value" id="stat-amount">$0</div>
        <div class="stat-label">销售金额</div>
      </div>
      <div class="stat-item highlight">
        <div class="stat-value" id="stat-commission">$0</div>
        <div class="stat-label">佣金</div>
      </div>
    </div>
    
    <!-- 商品表格 -->
    <table class="products-table">
      <thead>
        <tr>
          <th>ASIN</th>
          <th>商品名称</th>
          <th class="num">点击</th>
          <th class="num">详情浏览</th>
          <th class="num">加购</th>
          <th class="num">购买</th>
          <th class="num">销售金额</th>
          <th class="num">佣金</th>
        </tr>
      </thead>
      <tbody id="products-tbody">
      </tbody>
    </table>
  </div>

  <!-- Google Ads 花费 & ROI 板块 -->
  <div class="roi-card" id="roi-card" style="display:none;">
    <h3>💰 Google Ads 花费 & ROI <span style="font-size:12px;color:#888;font-weight:normal;">(汇率: 1 USD = <span id="exchange-rate">7.20</span> CNY)</span></h3>
    <div class="roi-content">
      <div class="roi-input-section">
        <label>Google Ads 花费 (￥)</label>
        <input type="number" id="ads-cost" value="720" step="0.01" min="0" onchange="calculateROI(); saveAdsCost()" onclick="this.select()">
        <span class="roi-hint">点击修改金额，自动保存</span>
      </div>
      <div class="roi-result-section">
        <div class="roi-item">
          <div class="roi-label">佣金收入 (￥)</div>
          <div class="roi-value" id="roi-commission">￥0</div>
        </div>
        <div class="roi-item">
          <div class="roi-label">广告花费 (￥)</div>
          <div class="roi-value" id="roi-cost">￥720</div>
        </div>
        <div class="roi-item highlight">
          <div class="roi-label">ROI</div>
          <div class="roi-value" id="roi-value">0%</div>
        </div>
        <div class="roi-item">
          <div class="roi-label">利润 (￥)</div>
          <div class="roi-value" id="roi-profit">￥-720</div>
        </div>
      </div>
    </div>
  </div>

  <!-- AI分析板块 -->
  <div class="ai-analysis-card">
    <h3>🤖 AI 广告优化分析</h3>
    <div class="upload-row">
      <div class="upload-item" onclick="document.getElementById('file-keywords').click()">
        <input type="file" id="file-keywords" accept=".xlsx,.xls,.csv" onchange="handleUpload('keywords', event)">
        <div class="upload-item-text">📊 搜索关键字报告</div>
        <div class="upload-item-status" id="status-keywords">点击上传 .xlsx/.csv</div>
      </div>
      <div class="upload-item" onclick="document.getElementById('file-search_terms').click()">
        <input type="file" id="file-search_terms" accept=".xlsx,.xls,.csv" onchange="handleUpload('search_terms', event)">
        <div class="upload-item-text">🔍 搜索字词报告</div>
        <div class="upload-item-status" id="status-search_terms">点击上传 .xlsx/.csv</div>
      </div>
      <button class="btn btn-success" onclick="startAnalysis()" id="start-btn">🚀 开始AI分析</button>
    </div>
    
    <div class="analysis-result" id="analysis-result" style="display:none;"></div>
    <div id="download-area" style="display:none; margin-top:16px;">
      <button class="btn btn-primary" onclick="downloadReport()">📥 下载优化报告 (TXT)</button>
    </div>
  </div>
</div>

<script>
// 全局变量
const uploadedFiles = { keywords: null, search_terms: null };
let currentMerchantId = null;
let currentReportFile = null;

// 页面初始化
document.addEventListener('DOMContentLoaded', function() {
  initDateInputs();
  loadMerchants();
  loadTotalAdsCost();
});

// 初始化日期
function initDateInputs() {
  const today = new Date();
  document.getElementById('date-start').value = '2026-01-01';
  document.getElementById('date-end').value = today.toISOString().split('T')[0];
}

// 快捷日期
function setDateRange(range) {
  const today = new Date();
  let start, end;
  if (range === 'this_year') {
    start = new Date(today.getFullYear(), 0, 1);  // 今年1月1日
    end = today;
  } else if (range === 'this_month') {
    start = new Date(today.getFullYear(), today.getMonth(), 1);
    end = today;
  } else if (range === 'last_month') {
    start = new Date(today.getFullYear(), today.getMonth() - 1, 1);
    end = new Date(today.getFullYear(), today.getMonth(), 0);
  } else if (range === 'last_7days') {
    start = new Date(today.getTime() - 6 * 24 * 60 * 60 * 1000);
    end = today;
  }
  document.getElementById('date-start').value = start.toISOString().split('T')[0];
  document.getElementById('date-end').value = end.toISOString().split('T')[0];
}

// 下载 Excel
function downloadExcel() {
  const selectId = document.getElementById('merchant-select').value;
  const inputId = document.getElementById('merchant-id-input').value.trim();
  const merchantId = selectId || inputId;
  const startDate = document.getElementById('date-start').value;
  const endDate = document.getElementById('date-end').value;
  
  if (!merchantId) {
    alert('请先选择商户');
    return;
  }
  
  const url = '/api/optimize/download_excel/' + merchantId + '?start_date=' + startDate + '&end_date=' + endDate;
  window.location.href = url;
}

// 加载商户列表
async function loadMerchants() {
  console.log('[DEBUG] loadMerchants called');
  try {
    const response = await fetch('/api/optimize/merchants');
    const result = await response.json();
    console.log('[DEBUG] API result:', result);
    if (result.success && result.merchants) {
      const select = document.getElementById('merchant-select');
      console.log('[DEBUG] select element:', select);
      console.log('[DEBUG] merchants count:', result.merchants.length);
      result.merchants.forEach(m => {
        const option = document.createElement('option');
        option.value = m.merchant_id;
        option.textContent = m.merchant_name;
        select.appendChild(option);
      });
      console.log('[DEBUG] options added');
    }
  } catch (e) {
    console.error('加载商户列表失败:', e);
  }
}

// 从下拉框选择商户
function onMerchantSelect() {
  const merchantId = document.getElementById('merchant-select').value;
  if (merchantId) {
    document.getElementById('merchant-id-input').value = ''; // 清空输入框
  }
  updateMerchantId();
}

// 从输入框输入商户ID
function onMerchantInput() {
  const inputId = document.getElementById('merchant-id-input').value.trim();
  if (inputId) {
    document.getElementById('merchant-select').value = ''; // 清空下拉框
  }
  updateMerchantId();
}

// 更新当前商户ID
function updateMerchantId() {
  const selectId = document.getElementById('merchant-select').value;
  const inputId = document.getElementById('merchant-id-input').value.trim();
  currentMerchantId = selectId || inputId || null;
  
  document.getElementById('fetch-btn').disabled = !currentMerchantId;
  
  if (currentMerchantId) {
    loadSalesData();
  } else {
    document.getElementById('sales-card').style.display = 'none';
  }
}

// 加载已有销售数据
async function loadSalesData() {
  const selectId = document.getElementById('merchant-select').value;
  const inputId = document.getElementById('merchant-id-input').value.trim();
  const merchantId = selectId || inputId;
  const startDate = document.getElementById('date-start').value;
  const endDate = document.getElementById('date-end').value;
  
  if (!merchantId) return;
  
  document.getElementById('fetch-status').textContent = '正在加载数据...';
  
  try {
    const url = '/api/optimize/yp_data/' + merchantId + '?start_date=' + startDate + '&end_date=' + endDate;
    const response = await fetch(url);
    const result = await response.json();
    
    if (result.success && result.data) {
      displaySalesData(result.data, result.products || []);
      if (result.data.clicks || result.data.commission) {
        document.getElementById('fetch-status').innerHTML = '<span style="color:#4caf50;">✓ 已从数据库加载</span> <button class="btn btn-secondary btn-sm" onclick="confirmRefresh()" style="margin-left:8px;">🔄 重新拉取</button>';
      } else {
        document.getElementById('fetch-status').innerHTML = '<span style="color:#ffa726;">暂无数据，请点击"拉取数据"</span>';
      }
    }
  } catch (e) {
    document.getElementById('fetch-status').innerHTML = '<span style="color:#ef5350;">加载失败: ' + e.message + '</span>';
  }
}

// 显示销售数据
function displaySalesData(summary, products) {
  document.getElementById('sales-card').style.display = 'block';
  document.getElementById('sales-date-range').textContent = '(' + document.getElementById('date-start').value + ' ~ ' + document.getElementById('date-end').value + ')';
  
  // 汇总统计
  document.getElementById('stat-clicks').textContent = (summary.clicks || 0).toLocaleString();
  document.getElementById('stat-views').textContent = (summary.detail_views || 0).toLocaleString();
  document.getElementById('stat-carts').textContent = (summary.add_to_carts || 0).toLocaleString();
  document.getElementById('stat-purchases').textContent = (summary.purchases || 0).toLocaleString();
  document.getElementById('stat-amount').textContent = '$' + (summary.amount || 0).toLocaleString();
  document.getElementById('stat-commission').textContent = '$' + (summary.commission || 0).toLocaleString();
  
  // 保存商户ID和佣金数据用于 ROI 计算
  window.currentMerchantId = summary.merchant_id;
  window.currentCommission = summary.commission || 0;
  
  // 显示 ROI 板块并加载已保存的广告花费
  document.getElementById('roi-card').style.display = 'block';
  loadAdsCost();
  
  // 商品表格
  const tbody = document.getElementById('products-tbody');
  tbody.innerHTML = '';
  
  if (products && products.length > 0) {
    products.forEach(p => {
      const tr = document.createElement('tr');
      tr.innerHTML = `
        <td class="asin">${p.asin || '-'}</td>
        <td>${p.product_name || '-'}</td>
        <td class="num">${(p.clicks || 0).toLocaleString()}</td>
        <td class="num">${(p.detail_views || 0).toLocaleString()}</td>
        <td class="num">${(p.add_to_carts || 0).toLocaleString()}</td>
        <td class="num">${(p.purchases || 0).toLocaleString()}</td>
        <td class="num">$${(p.amount || 0).toLocaleString()}</td>
        <td class="num">$${(p.commission || 0).toLocaleString()}</td>
      `;
      tbody.appendChild(tr);
    });
  } else {
    tbody.innerHTML = '<tr><td colspan="8" style="text-align:center;color:#666;padding:40px;">暂无商品数据</td></tr>';
  }
}

// 确认重新拉取
function confirmRefresh() {
  const startDate = document.getElementById('date-start').value;
  const endDate = document.getElementById('date-end').value;
  const msg = '确定要重新拉取 ' + startDate + ' ~ ' + endDate + ' 的数据吗？\\n\\n这将覆盖数据库中的现有数据。';
  if (confirm(msg)) {
    fetchSalesData(true);
  }
}

// 拉取销售数据
async function fetchSalesData(force = false) {
  const selectId = document.getElementById('merchant-select').value;
  const inputId = document.getElementById('merchant-id-input').value.trim();
  const merchantId = selectId || inputId;
  const startDate = document.getElementById('date-start').value;
  const endDate = document.getElementById('date-end').value;
  
  if (!merchantId) return;
  
  // 检查是否已有数据
  const hasData = document.getElementById('stat-commission').textContent !== '$0';
  if (!force && hasData) {
    confirmRefresh();
    return;
  }
  
  document.getElementById('fetch-btn').disabled = true;
  document.getElementById('fetch-status').textContent = '正在从YP平台拉取数据...';
  
  try {
    const url = '/api/optimize/yp_data/' + merchantId + '?force=true&start_date=' + startDate + '&end_date=' + endDate;
    const response = await fetch(url, { method: 'POST' });
    const result = await response.json();
    
    if (result.success) {
      if (result.collecting) {
        document.getElementById('fetch-status').innerHTML = '<span style="color:#ffa726;">⏳ ' + result.message + '</span>';
        setTimeout(loadSalesData, 3000);
      } else {
        loadSalesData();
      }
    } else {
      document.getElementById('fetch-status').innerHTML = '<span style="color:#ef5350;">拉取失败: ' + (result.error || '未知错误') + '</span>';
    }
  } catch (e) {
    document.getElementById('fetch-status').innerHTML = '<span style="color:#ef5350;">请求失败: ' + e.message + '</span>';
  }
  
  document.getElementById('fetch-btn').disabled = false;
}

// 汇率 (1 USD = 7.20 CNY)
const EXCHANGE_RATE = 7.20;

// 计算 ROI
function calculateROI() {
  const costCNY = parseFloat(document.getElementById('ads-cost').value) || 0;
  const commissionUSD = window.currentCommission || 0;
  
  // 将佣金从美元换算成人民币
  const commissionCNY = commissionUSD * EXCHANGE_RATE;
  
  // 计算利润
  const profit = commissionCNY - costCNY;
  
  // 计算 ROI (利润 / 成本 * 100%)
  const roi = costCNY > 0 ? ((profit / costCNY) * 100).toFixed(1) : 0;
  
  // 更新显示
  document.getElementById('roi-commission').textContent = '￥' + commissionCNY.toLocaleString(undefined, {minimumFractionDigits: 2, maximumFractionDigits: 2});
  document.getElementById('roi-cost').textContent = '￥' + costCNY.toLocaleString(undefined, {minimumFractionDigits: 2, maximumFractionDigits: 2});
  document.getElementById('roi-value').textContent = roi + '%';
  document.getElementById('roi-profit').textContent = (profit >= 0 ? '+' : '') + '￥' + profit.toLocaleString(undefined, {minimumFractionDigits: 2, maximumFractionDigits: 2});
  
  // 利润为负时显示红色
  const profitItem = document.getElementById('roi-profit').parentElement;
  if (profit < 0) {
    profitItem.classList.add('negative');
  } else {
    profitItem.classList.remove('negative');
  }
}

// 保存广告花费到数据库
async function saveAdsCost() {
  const merchantId = window.currentMerchantId;
  const costCNY = parseFloat(document.getElementById('ads-cost').value) || 0;
  
  if (!merchantId) return;
  
  try {
    const response = await fetch('/api/optimize/save_ads_cost', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ merchant_id: merchantId, ads_cost_cny: costCNY })
    });
    const result = await response.json();
    if (result.success) {
      console.log('广告花费已保存');
      loadTotalAdsCost();  // 刷新总花费
    }
  } catch (e) {
    console.error('保存广告花费失败:', e);
  }
}

// 加载已保存的广告花费
async function loadAdsCost() {
  const merchantId = window.currentMerchantId;
  
  if (!merchantId) {
    calculateROI();
    return;
  }
  
  try {
    const response = await fetch('/api/optimize/get_ads_cost/' + merchantId);
    const result = await response.json();
    if (result.success && result.ads_cost_cny !== undefined) {
      document.getElementById('ads-cost').value = result.ads_cost_cny;
    } else {
      // 默认值 720 元 (约 100 美元)
      document.getElementById('ads-cost').value = 720;
    }
    calculateROI();
  } catch (e) {
    console.error('加载广告花费失败:', e);
    document.getElementById('ads-cost').value = 720;
    calculateROI();
  }
}

// 加载所有品牌的 Google Ads 总花费
async function loadTotalAdsCost() {
  try {
    const response = await fetch('/api/optimize/total_ads_cost');
    const result = await response.json();
    if (result.success) {
      const totalCost = result.total_ads_cost || 0;
      const totalCommission = result.total_commission || 0;
      const totalProfit = result.total_profit || 0;
      
      document.getElementById('total-ads-cost').textContent = '￥' + totalCost.toLocaleString(undefined, {minimumFractionDigits: 2, maximumFractionDigits: 2});
      document.getElementById('total-commission').textContent = '￥' + totalCommission.toLocaleString(undefined, {minimumFractionDigits: 2, maximumFractionDigits: 2});
      
      const profitElement = document.getElementById('total-profit');
      profitElement.textContent = (totalProfit >= 0 ? '+' : '') + '￥' + totalProfit.toLocaleString(undefined, {minimumFractionDigits: 2, maximumFractionDigits: 2});
      
      // 利润为负时显示红色
      const profitItem = profitElement.parentElement;
      if (totalProfit < 0) {
        profitItem.classList.add('negative');
        profitItem.classList.remove('highlight');
      } else {
        profitItem.classList.remove('negative');
        profitItem.classList.add('highlight');
      }
    }
  } catch (e) {
    console.error('加载总花费失败:', e);
  }
}

// 文件上传
function handleUpload(type, event) {
  const file = event.target.files[0];
  if (!file) return;
  uploadedFiles[type] = file;
  document.getElementById('upload-' + type).classList.add('has-file');
  document.getElementById('status-' + type).textContent = '✓ ' + file.name;
}

// 开始AI分析
async function startAnalysis() {
  const costCNY = parseFloat(document.getElementById('ads-cost').value) || 0;
  const costUSD = (costCNY / EXCHANGE_RATE).toFixed(2);  // 换算成美元
  const merchantSelect = document.getElementById('merchant-select');
  const inputId = document.getElementById('merchant-id-input').value.trim();
  
  // 获取品牌名称：优先下拉框选中项，其次输入的商户ID
  let brandName = 'Unknown';
  if (merchantSelect.value && merchantSelect.selectedIndex > 0) {
    brandName = merchantSelect.options[merchantSelect.selectedIndex].text;
  } else if (inputId) {
    brandName = '商户' + inputId;
  }
  
  if (!uploadedFiles.keywords && !uploadedFiles.search_terms) {
    alert('请至少上传一份报告文件');
    return;
  }
  
  document.getElementById('start-btn').disabled = true;
  document.getElementById('analysis-result').style.display = 'block';
  document.getElementById('analysis-result').textContent = '正在分析中，请稍候...';
  
  const formData = new FormData();
  if (uploadedFiles.keywords) formData.append('keywords_file', uploadedFiles.keywords);
  if (uploadedFiles.search_terms) formData.append('search_terms_file', uploadedFiles.search_terms);
  formData.append('yp_clicks', document.getElementById('stat-clicks').textContent.replace(/,/g, '') || 0);
  formData.append('yp_add_to_carts', document.getElementById('stat-carts').textContent.replace(/,/g, '') || 0);
  formData.append('yp_purchases', document.getElementById('stat-purchases').textContent.replace(/,/g, '') || 0);
  formData.append('yp_commission', document.getElementById('stat-commission').textContent.replace('$', '').replace(/,/g, '') || 0);
  formData.append('cost_amount', costUSD);  // 传美元
  formData.append('cost_amount_cny', costCNY);  // 同时传人民币
  formData.append('brand_name', brandName);
  
  try {
    const response = await fetch('/api/optimize/agent', { method: 'POST', body: formData });
    const result = await response.json();
    
    if (result.success) {
      document.getElementById('analysis-result').textContent = result.report;
      currentReportFile = result.filename;
      document.getElementById('download-area').style.display = 'block';
    } else {
      document.getElementById('analysis-result').textContent = '❌ 分析失败：' + (result.error || '未知错误');
    }
  } catch (e) {
    document.getElementById('analysis-result').textContent = '❌ 请求失败：' + e.message;
  }
  
  document.getElementById('start-btn').disabled = false;
}

// 下载报告
function downloadReport() {
  if (currentReportFile) {
    window.location.href = '/api/optimize/download/' + currentReportFile;
  }
}
</script>
</body>
</html>
"""
)


@bp.route("/optimize")
def optimize_page():
    return render_template_string(OPTIMIZE_HTML)


# ═══════════════════════════════════════════════════════════════════════════
# Agent 优化 API
# ═══════════════════════════════════════════════════════════════════════════


@bp.route("/api/optimize/agent", methods=["POST"])
def api_optimize_agent():
    """Agent模式优化API - 接收文件和数据，调用google-ads-optimizer技能"""
    try:
        import os
        import json
        from datetime import datetime

        # 获取上传的文件
        keywords_file = request.files.get("keywords_file")
        search_terms_file = request.files.get("search_terms_file")

        # 获取手动输入的数据
        yp_clicks = int(request.form.get("yp_clicks", 0) or 0)
        yp_add_to_carts = int(request.form.get("yp_add_to_carts", 0) or 0)
        yp_purchases = int(request.form.get("yp_purchases", 0) or 0)
        yp_commission = float(request.form.get("yp_commission", 0) or 0)
        cost_amount = float(request.form.get("cost_amount", 0) or 0)
        brand_name = request.form.get("brand_name", "Unknown")

        # 解析Excel文件
        keywords_data = []
        search_terms_data = []

        def parse_excel(file):
            """解析Excel文件"""
            if not file:
                return []
            rows = []
            try:
                import openpyxl

                file.seek(0)
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                headers = [
                    str(cell.value).strip() if cell.value else "" for cell in ws[1]
                ]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = {}
                    for idx, val in enumerate(row):
                        if idx < len(headers) and headers[idx]:
                            row_dict[headers[idx]] = val
                    if row_dict:
                        rows.append(row_dict)
            except Exception as e:
                print(f"[ERROR] 解析Excel失败: {e}")
            return rows

        keywords_data = parse_excel(keywords_file)
        search_terms_data = parse_excel(search_terms_file)

        # 计算基础指标
        total_cost = sum(
            float(row.get("Cost", 0) or row.get("花费", 0) or 0)
            for row in keywords_data
        )
        total_clicks = sum(
            int(row.get("Clicks", 0) or row.get("点击", 0) or 0)
            for row in keywords_data
        )
        total_impressions = sum(
            int(row.get("Impr.", 0) or row.get("展示", 0) or 0) for row in keywords_data
        )

        # 使用手动输入的花费（如果有）
        if cost_amount > 0:
            total_cost = cost_amount

        # 计算ROI
        roi = (yp_commission - total_cost) / total_cost if total_cost > 0 else 0
        roas = yp_commission / total_cost if total_cost > 0 else 0

        # 分析关键词
        good_keywords = []
        optimize_keywords = []
        pause_keywords = []
        negative_terms = []

        for kw in keywords_data:
            keyword = (
                kw.get("Keyword") or kw.get("关键字") or kw.get("Search keyword", "")
            )
            clicks = int(kw.get("Clicks", 0) or kw.get("点击", 0) or 0)
            cost = float(kw.get("Cost", 0) or kw.get("花费", 0) or 0)
            conversions = int(kw.get("Conversions", 0) or kw.get("转化", 0) or 0)
            match_type = kw.get("Match type") or kw.get("匹配类型", "")
            cpc = float(kw.get("Avg. CPC", 0) or kw.get("平均CPC", 0) or 0)

            if not keyword:
                continue

            # 效果评分
            score = (conversions * 10) + (clicks * 0.5) - (cost * 0.1)

            if score > 5:
                good_keywords.append(
                    {
                        "keyword": keyword,
                        "match_type": match_type,
                        "clicks": clicks,
                        "cost": cost,
                        "conversions": conversions,
                        "cpc": cpc,
                    }
                )
            elif cost > 5 and clicks < 3:
                pause_keywords.append(
                    {
                        "keyword": keyword,
                        "cost": cost,
                        "clicks": clicks,
                        "reason": "花费高但点击少",
                    }
                )
            elif clicks > 10 and conversions == 0:
                optimize_keywords.append(
                    {
                        "keyword": keyword,
                        "match_type": match_type,
                        "clicks": clicks,
                        "cost": cost,
                        "suggestion": "考虑收紧匹配类型或调整出价",
                    }
                )

        # 分析搜索字词，提取否定词候选
        for term in search_terms_data:
            search_term = (
                term.get("Search term")
                or term.get("搜索词")
                or term.get("Search query", "")
            )
            clicks = int(term.get("Clicks", 0) or term.get("点击", 0) or 0)
            cost = float(term.get("Cost", 0) or term.get("花费", 0) or 0)
            conversions = int(term.get("Conversions", 0) or term.get("转化", 0) or 0)

            if not search_term:
                continue

            # 检查是否是低意图词
            low_intent_words = [
                "free",
                "cheap",
                "review",
                "how to",
                "tutorial",
                "walmart",
                "target",
                "refurbished",
            ]
            if any(word in search_term.lower() for word in low_intent_words):
                if clicks > 3 or cost > 2:
                    negative_terms.append(
                        {
                            "term": search_term,
                            "clicks": clicks,
                            "cost": cost,
                            "reason": "低意图搜索词",
                        }
                    )
            elif clicks > 5 and cost > 3 and conversions == 0:
                negative_terms.append(
                    {
                        "term": search_term,
                        "clicks": clicks,
                        "cost": cost,
                        "reason": "高花费无转化",
                    }
                )

        # 生成报告
        report_date = datetime.now().strftime("%Y-%m-%d")
        report_lines = [
            "═" * 60,
            "📊 广告优化报告",
            f"品牌: {brand_name}",
            f"日期: {report_date}",
            "═" * 60,
            "",
            "【整体效果】",
            f"  Google Ads 花费: ${total_cost:.2f}",
            f"  YP 佣金: ${yp_commission:.2f}",
            f"  ROI: {roi * 100:.1f}%",
            f"  ROAS: {roas:.2f}x",
            "",
            "【数据来源】",
            f"  搜索关键字: {len(keywords_data)} 条",
            f"  搜索字词: {len(search_terms_data)} 条",
            f"  YP 点击: {yp_clicks}",
            f"  YP 加购: {yp_add_to_carts}",
            f"  YP 购买: {yp_purchases}",
            "",
            "【优化建议摘要】",
            f"  ✅ 优质关键词: {len(good_keywords)} 个",
            f"  🔧 需优化关键词: {len(optimize_keywords)} 个",
            f"  🚫 建议否定词: {len(negative_terms)} 个",
            f"  ⏸️ 建议暂停词: {len(pause_keywords)} 个",
            "",
            "─" * 60,
            "✅ 优质关键词（继续投放）",
            "─" * 60,
        ]

        if good_keywords:
            for i, kw in enumerate(good_keywords[:10], 1):
                report_lines.append(
                    f"{i}. {kw['keyword']} | 匹配: {kw['match_type']} | "
                    f"点击: {kw['clicks']} | 花费: ${kw['cost']:.2f} | 转化: {kw['conversions']}"
                )
        else:
            report_lines.append("  暂无优质关键词")

        report_lines.extend(
            [
                "",
                "─" * 60,
                "🔧 需优化关键词",
                "─" * 60,
            ]
        )

        if optimize_keywords:
            for i, kw in enumerate(optimize_keywords[:10], 1):
                report_lines.append(f"{i}. {kw['keyword']}")
                report_lines.append(
                    f"   - 点击: {kw['clicks']} | 花费: ${kw['cost']:.2f}"
                )
                report_lines.append(f"   - 建议: {kw['suggestion']}")
        else:
            report_lines.append("  暂无需优化的关键词")

        report_lines.extend(
            [
                "",
                "─" * 60,
                "🚫 建议添加否定词",
                "─" * 60,
            ]
        )

        if negative_terms:
            for i, term in enumerate(negative_terms[:15], 1):
                report_lines.append(
                    f'{i}. "{term["term"]}" - 点击: {term["clicks"]} | 花费: ${term["cost"]:.2f} | {term["reason"]}'
                )
        else:
            report_lines.append("  暂无建议否定词")

        report_lines.extend(
            [
                "",
                "─" * 60,
                "⏸️ 建议暂停关键词",
                "─" * 60,
            ]
        )

        if pause_keywords:
            for i, kw in enumerate(pause_keywords[:10], 1):
                report_lines.append(
                    f"{i}. {kw['keyword']} - 花费: ${kw['cost']:.2f} | 点击: {kw['clicks']} | {kw['reason']}"
                )
        else:
            report_lines.append("  暂无建议暂停的关键词")

        # 添加出价建议
        report_lines.extend(
            [
                "",
                "─" * 60,
                "💰 出价调整建议",
                "─" * 60,
            ]
        )

        if roi > 0.5:
            report_lines.append("  📈 ROI > 50%，建议加码投放，可提高出价 10-20%")
        elif roi > 0:
            report_lines.append("  ➡️ ROI 为正但较低，建议维持当前出价，优化关键词质量")
        elif roi > -0.3:
            report_lines.append("  📉 ROI 为负，建议降低出价 20-30%，暂停亏损关键词")
        else:
            report_lines.append("  ⚠️ ROI 严重亏损，建议大幅降低预算或暂停投放")

        report_lines.extend(
            [
                "",
                "═" * 60,
                "📅 下次优化建议",
                "═" * 60,
                f"1. 建议在 1-2 周后再次上传数据进行对比分析",
                f"2. 优先处理 {len(negative_terms)} 个否定词，减少无效花费",
                f"3. 关注 {len(pause_keywords)} 个亏损关键词，及时止损",
                "",
                "报告生成时间: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ]
        )

        report_content = "\n".join(report_lines)

        # 保存报告文件
        temp_dir = BASE_DIR / "temp"
        temp_dir.mkdir(exist_ok=True)
        filename = f"优化报告_{brand_name}_{report_date}.txt"
        filepath = temp_dir / filename

        with open(filepath, "w", encoding="utf-8") as f:
            f.write(report_content)

        return jsonify(
            {
                "success": True,
                "report": report_content,
                "filename": filename,
                "summary": {
                    "roi": roi,
                    "roas": roas,
                    "good_keywords": len(good_keywords),
                    "optimize_keywords": len(optimize_keywords),
                    "negative_terms": len(negative_terms),
                    "pause_keywords": len(pause_keywords),
                },
            }
        )

    except Exception as e:
        import traceback

        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


@bp.route("/api/optimize/download/<filename>")
def api_optimize_download(filename):
    """下载优化报告"""
    try:
        from flask import send_file

        filepath = BASE_DIR / "temp" / filename
        if filepath.exists():
            return send_file(filepath, as_attachment=True, download_name=filename)
        else:
            return jsonify({"success": False, "error": "文件不存在"}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════
# T-012: QS 质量评分仪表盘
# ═══════════════════════════════════════════════════════════════════════════

QS_DASHBOARD_HTML = (
    """
<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>质量评分仪表盘 - YP Affiliate</title>
<style>
"""
    + BASE_CSS
    + """
/* QS仪表盘专用样式 */
.qs-stats-grid {
  display: grid; grid-template-columns: repeat(auto-fit, minmax(240px, 1fr));
  gap: 20px; margin-bottom: 28px;
}
.qs-stat-card {
  background: #1a1d24; border: 1px solid #2a2d36; border-radius: 12px;
  padding: 24px; text-align: center;
}
.qs-stat-value {
  font-size: 36px; font-weight: 700; margin-bottom: 8px;
}
.qs-stat-value.good { color: #66bb6a; }
.qs-stat-value.warning { color: #ffa726; }
.qs-stat-value.danger { color: #ef5350; }
.qs-stat-label { font-size: 13px; color: #888; }

/* QS徽章 */
.qs-badge-circle {
  display: inline-flex; align-items: center; justify-content: center;
  width: 44px; height: 44px; border-radius: 50%; font-size: 14px; font-weight: 700;
}
.qs-badge-circle.good { background: rgba(46, 125, 50, 0.2); color: #66bb6a; border: 2px solid #2e7d32; }
.qs-badge-circle.warning { background: rgba(255, 167, 38, 0.2); color: #ffa726; border: 2px solid #ffa726; }
.qs-badge-circle.danger { background: rgba(198, 40, 40, 0.2); color: #ef5350; border: 2px solid #c62828; }

/* 商品列表 */
.product-list { background: #1a1d24; border: 1px solid #2a2d36; border-radius: 12px; overflow: hidden; }
.product-list th { background: #15181f; }
.product-list td { vertical-align: middle; }
.product-name { max-width: 300px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
.product-row:hover td { background: #1e2129; }

/* 进度条 */
.progress-bar-mini {
  height: 6px; background: #23262f; border-radius: 3px; overflow: hidden; width: 100px;
}
.progress-fill-mini {
  height: 100%; border-radius: 3px; transition: width .3s ease;
}
.progress-fill-mini.good { background: #2e7d32; }
.progress-fill-mini.warning { background: #ffa726; }
.progress-fill-mini.danger { background: #c62828; }

/* 空状态 */
.empty-state { text-align: center; padding: 60px 20px; color: #666; }
.empty-state-icon { font-size: 64px; margin-bottom: 16px; }
</style>
</head>
<body>
"""
    + NAV_HTML.format(
        p0="",
        p1="",
        p2="",
        p3="",
        p4="",
        p5="",
        p6="",
        p7="active",
        p8="",
        p9="",
        p10="",
        p11="",
    )
    + """
<div class="container">
  <h1 style="font-size: 20px; margin-bottom: 24px;">⭐ 质量评分仪表盘</h1>
  
  <!-- 总览卡片 -->
  <div class="qs-stats-grid" id="qs-stats">
    <div class="qs-stat-card">
      <div class="qs-stat-value" id="stat-avg">-</div>
      <div class="qs-stat-label">平均 QS 分数</div>
    </div>
    <div class="qs-stat-card">
      <div class="qs-stat-value good" id="stat-good">-</div>
      <div class="qs-stat-label">QS ≥ 80 的广告</div>
    </div>
    <div class="qs-stat-card">
      <div class="qs-stat-value danger" id="stat-bad">-</div>
      <div class="qs-stat-label">QS < 60 的广告</div>
    </div>
    <div class="qs-stat-card">
      <div class="qs-stat-value" id="stat-total">-</div>
      <div class="qs-stat-label">总广告数</div>
    </div>
  </div>
  
  <!-- 商品列表 -->
  <div class="card">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px;">
      <h2 style="margin:0;font-size:15px;">商品 QS 排名（按平均 QS 升序）</h2>
      <button class="btn btn-primary btn-sm" onclick="loadQsData()">🔄 刷新数据</button>
    </div>
    <table class="product-list">
      <thead>
        <tr>
          <th>ASIN</th>
          <th>商品名称</th>
          <th>平均 QS</th>
          <th>广告数量</th>
          <th>QS 分布</th>
          <th>操作</th>
        </tr>
      </thead>
      <tbody id="product-tbody">
        <tr><td colspan="6" style="text-align:center;padding:40px;color:#666;">加载中...</td></tr>
      </tbody>
    </table>
  </div>
</div>

<script>
function toast(msg, type) {
  type = type || 'info';
  var c = document.getElementById('toast-container');
  var t = document.createElement('div');
  t.className = 'toast toast-' + type;
  t.textContent = msg;
  c.appendChild(t);
  setTimeout(function() { t.remove(); }, 3500);
}

function getQsClass(score) {
  if (score >= 80) return 'good';
  if (score >= 60) return 'warning';
  return 'danger';
}

function htmlEsc(s) {
  return String(s || '').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;').replace(/'/g,'&#39;');
}

async function loadQsData() {
  try {
    const res = await fetch('/api/qs/dashboard');
    const data = await res.json();
    
    if (!data.ok) {
      toast(data.msg || '加载失败', 'error');
      return;
    }
    
    // 更新统计卡片
    document.getElementById('stat-avg').textContent = data.avg_qs || 0;
    document.getElementById('stat-avg').className = 'qs-stat-value ' + getQsClass(data.avg_qs || 0);
    document.getElementById('stat-good').textContent = data.qs_good_count || 0;
    document.getElementById('stat-bad').textContent = data.qs_bad_count || 0;
    document.getElementById('stat-total').textContent = data.total_ads || 0;
    
    // 更新商品列表（后端返回字段：items / avg_quality_score / qs_good / qs_bad / scored_ads）
    const tbody = document.getElementById('product-tbody');
    if (!data.items || data.items.length === 0) {
      tbody.innerHTML = '<tr><td colspan="6" class="empty-state"><div class="empty-state-icon">📊</div><div>暂无数据</div></td></tr>';
      return;
    }
    
    try {
      tbody.innerHTML = data.items.map(p => {
        const avgQs = p.avg_quality_score || 0;
        const qsClass = getQsClass(avgQs);
        const adCount = p.scored_ads || p.ad_count || 0;
        const goodPct = adCount > 0 ? Math.round((p.qs_good || 0) / adCount * 100) : 0;
        const badPct  = adCount > 0 ? Math.round((p.qs_bad  || 0) / adCount * 100) : 0;
        const avgQsDisplay = avgQs ? Math.round(avgQs) : '-';
        const safeName = htmlEsc(p.product_name);
        const safeAsin = htmlEsc(p.asin);
        
        return `
          <tr class="product-row">
            <td><code>${safeAsin}</code></td>
            <td class="product-name" title="${safeName}">${safeName || '-'}</td>
            <td>
              <span class="qs-badge-circle ${avgQs ? qsClass : 'warning'}">${avgQsDisplay}</span>
            </td>
            <td>${adCount}</td>
            <td>
              <div style="display:flex;align-items:center;gap:8px;font-size:12px;">
                <span style="color:#66bb6a">${p.qs_good || 0} 优</span>
                <div class="progress-bar-mini">
                  <div class="progress-fill-mini good" style="width:${goodPct}%"></div>
                </div>
                <span style="color:#ef5350">${p.qs_bad || 0} 差</span>
                <div class="progress-bar-mini">
                  <div class="progress-fill-mini danger" style="width:${badPct}%"></div>
                </div>
              </div>
            </td>
            <td>
              <a href="/plans/${safeAsin}" class="btn btn-secondary btn-sm">查看</a>
              <button class="btn btn-primary btn-sm" data-asin="${safeAsin}" onclick="rescoreProduct(this.dataset.asin)">重评</button>
            </td>
          </tr>
        `;
      }).join('');
    } catch(renderErr) {
      console.error('渲染商品列表失败:', renderErr);
      toast('渲染失败: ' + renderErr.message, 'error');
    }
    
  } catch (e) {
    console.error('loadQsData失败:', e);
    toast(e.message, 'error');
  }
}

async function rescoreProduct(asin) {
  try {
    const res = await fetch('/api/ads/score/' + asin, { method: 'POST' });
    const data = await res.json();
    if (data.ok) {
      toast('评分完成！平均QS: ' + (data.avg_qs || 0), 'success');
      loadQsData();
    } else {
      toast(data.msg || '评分失败', 'error');
    }
  } catch (e) {
    toast(e.message, 'error');
  }
}

window.onload = loadQsData;
</script>
</body>
</html>
"""
)


@bp.route("/qs_dashboard")
def qs_dashboard_page():
    return render_template_string(QS_DASHBOARD_HTML)


# ═══════════════════════════════════════════════════════════════════════════
# T-013: 竞品文案参考库
# ═══════════════════════════════════════════════════════════════════════════

COMPETITOR_ADS_HTML = (
    """
<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>竞品文案参考库 - YP Affiliate</title>
<style>
"""
    + BASE_CSS
    + """
/* 竞品文案库专用样式 */
.competitor-layout {
  display: grid; grid-template-columns: 260px 1fr; gap: 20px;
}
.sidebar {
  background: #1a1d24; border: 1px solid #2a2d36; border-radius: 12px;
  padding: 16px; height: fit-content;
}
.sidebar h3 { font-size: 14px; color: #888; margin-bottom: 12px; text-transform: uppercase; }
.merchant-list { max-height: 400px; overflow-y: auto; }
.merchant-item {
  display: flex; align-items: center; gap: 8px; padding: 8px 10px;
  border-radius: 6px; cursor: pointer; transition: background .15s;
}
.merchant-item:hover { background: #23262f; }
.merchant-item.active { background: #1565c033; border-left: 3px solid #1565c0; }
.merchant-checkbox { width: 16px; height: 16px; accent-color: #1565c0; }
.merchant-name { font-size: 13px; color: #e0e0e0; }
.merchant-count { font-size: 11px; color: #888; margin-left: auto; }

/* 搜索栏 */
.search-bar {
  display: flex; gap: 12px; margin-bottom: 20px;
}
.search-input {
  flex: 1; padding: 10px 16px; border: 1px solid #2a2d36;
  border-radius: 8px; background: #15181f; color: #e0e0e0; font-size: 14px;
}
.search-input:focus { outline: none; border-color: #1565c0; }

/* 统计栏 */
.stats-bar-comp {
  display: flex; gap: 24px; padding: 16px 20px;
  background: #1a1d24; border-radius: 8px; margin-bottom: 20px;
}
.stat-comp { text-align: center; }
.stat-comp-value { font-size: 24px; font-weight: 700; color: #fff; }
.stat-comp-label { font-size: 12px; color: #888; margin-top: 4px; }

/* 文案卡片网格 */
.ads-grid {
  display: grid; grid-template-columns: repeat(auto-fill, minmax(360px, 1fr));
  gap: 16px;
}
.ad-card {
  background: #1a1d24; border: 1px solid #2a2d36; border-radius: 12px;
  padding: 20px; transition: border-color .15s;
}
.ad-card:hover { border-color: #1565c0; }
.ad-card-header {
  display: flex; justify-content: space-between; align-items: center;
  margin-bottom: 12px; padding-bottom: 12px; border-bottom: 1px solid #23262f;
}
.ad-card-source { font-size: 12px; color: #888; }
.ad-card-date { font-size: 11px; color: #666; }
.ad-card-headline {
  font-size: 15px; font-weight: 600; color: #64b5f6; margin-bottom: 10px;
  line-height: 1.4;
}
.ad-card-desc {
  font-size: 13px; color: #adb5bd; line-height: 1.6; margin-bottom: 12px;
}
.ad-card-footer {
  display: flex; gap: 8px;
}
.ad-card-tag {
  font-size: 11px; padding: 2px 8px; border-radius: 4px;
  background: #23262f; color: #888;
}

/* 复制按钮 */
.copy-btn {
  margin-left: auto; padding: 4px 12px; font-size: 12px;
  background: #23262f; border: 1px solid #2a2d36; color: #adb5bd;
  border-radius: 6px; cursor: pointer; transition: all .15s;
}
.copy-btn:hover { background: #1565c0; color: #fff; border-color: #1565c0; }
.copy-btn.copied { background: #2e7d32; color: #fff; border-color: #2e7d32; }

/* 空状态 */
.empty-state-comp { text-align: center; padding: 60px; color: #666; }
.empty-state-comp-icon { font-size: 64px; margin-bottom: 16px; }
</style>
</head>
<body>
"""
    + NAV_HTML.format(
        p0="",
        p1="",
        p2="",
        p3="",
        p4="",
        p5="",
        p6="",
        p7="",
        p8="active",
        p9="",
        p10="",
        p11="",
    )
    + """
<div class="container">
  <h1 style="font-size: 20px; margin-bottom: 20px;">🔍 竞品文案参考库</h1>
  
  <!-- 统计栏 -->
  <div class="stats-bar-comp" id="stats-bar">
    <div class="stat-comp">
      <div class="stat-comp-value" id="stat-total-ads">-</div>
      <div class="stat-comp-label">总文案数</div>
    </div>
    <div class="stat-comp">
      <div class="stat-comp-value" id="stat-merchants">-</div>
      <div class="stat-comp-label">商户数</div>
    </div>
    <div class="stat-comp">
      <div class="stat-comp-value" id="stat-latest" style="font-size:14px;color:#888;">-</div>
      <div class="stat-comp-label">最新采集时间</div>
    </div>
  </div>
  
  <div class="competitor-layout">
    <!-- 左侧筛选 -->
    <div class="sidebar">
      <h3>🏢 商户筛选</h3>
      <div style="margin-bottom:12px;">
        <button class="btn btn-secondary btn-sm" onclick="selectAllMerchants(true)">全选</button>
        <button class="btn btn-secondary btn-sm" onclick="selectAllMerchants(false)">清空</button>
      </div>
      <div class="merchant-list" id="merchant-list">
        <div style="color:#666;padding:20px;text-align:center;">加载中...</div>
      </div>
    </div>
    
    <!-- 右侧内容 -->
    <div>
      <!-- 搜索栏 -->
      <div class="search-bar">
        <input type="text" class="search-input" id="search-input" placeholder="搜索文案内容..." onkeyup="handleSearch()">
        <button class="btn btn-primary" onclick="loadCompetitorAds()">🔍 搜索</button>
        <button class="btn btn-secondary" onclick="clearSearch()">清空</button>
      </div>
      
      <!-- 文案网格 -->
      <div class="ads-grid" id="ads-grid">
        <div class="empty-state-comp">
          <div class="empty-state-comp-icon">📋</div>
          <div>请选择商户或输入关键词搜索</div>
        </div>
      </div>
    </div>
  </div>
</div>

<script>
let allMerchants = [];
let selectedMerchants = new Set();
let allAds = [];

function toast(msg, type) {
  type = type || 'info';
  var c = document.getElementById('toast-container');
  var t = document.createElement('div');
  t.className = 'toast toast-' + type;
  t.textContent = msg;
  c.appendChild(t);
  setTimeout(function() { t.remove(); }, 3500);
}

function compHtmlEsc(s) {
  return String(s || '').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;').replace(/'/g,'&#39;');
}

async function loadMerchants() {
  try {
    const res = await fetch('/api/competitor/merchants');
    const data = await res.json();
    
    if (!data.ok) {
      console.error('loadMerchants API error:', data.msg);
      document.getElementById('merchant-list').innerHTML = '<div style="color:#ef5350;padding:20px;text-align:center;">加载失败: ' + compHtmlEsc(data.msg || '未知错误') + '</div>';
      toast(data.msg || '加载商户失败', 'error');
      return;
    }
    
    allMerchants = data.merchants || [];
    renderMerchantList();
    updateStats(data.stats || {});
  } catch (e) {
    console.error('loadMerchants 异常:', e);
    document.getElementById('merchant-list').innerHTML = '<div style="color:#ef5350;padding:20px;text-align:center;">加载异常: ' + compHtmlEsc(e.message) + '</div>';
    toast(e.message, 'error');
  }
}

function renderMerchantList() {
  const container = document.getElementById('merchant-list');
  if (allMerchants.length === 0) {
    container.innerHTML = '<div style="color:#666;padding:20px;text-align:center;">暂无商户数据</div>';
    return;
  }
  
  try {
    // m.id 统一转字符串，避免数字/字符串 Set 判断失效
    container.innerHTML = allMerchants.map(m => {
      const sid = String(m.id);
      const isActive = selectedMerchants.has(sid);
      const safeName = compHtmlEsc(m.name);
      const safeId = compHtmlEsc(sid);
      return `<div class="merchant-item ${isActive ? 'active' : ''}" data-id="${safeId}" onclick="toggleMerchant(this.dataset.id)">
        <input type="checkbox" class="merchant-checkbox" ${isActive ? 'checked' : ''} onclick="event.stopPropagation()">
        <span class="merchant-name">${safeName}</span>
        <span class="merchant-count">${m.ad_count || 0}</span>
      </div>`;
    }).join('');
  } catch(renderErr) {
    console.error('renderMerchantList 渲染失败:', renderErr);
    container.innerHTML = '<div style="color:#ef5350;padding:20px;text-align:center;">渲染失败: ' + compHtmlEsc(renderErr.message) + '</div>';
  }
}

function toggleMerchant(id) {
  // id 统一转字符串
  const sid = String(id);
  if (selectedMerchants.has(sid)) {
    selectedMerchants.delete(sid);
  } else {
    selectedMerchants.add(sid);
  }
  renderMerchantList();
  loadCompetitorAds();
}

function selectAllMerchants(select) {
  if (select) {
    allMerchants.forEach(m => selectedMerchants.add(String(m.id)));
  } else {
    selectedMerchants.clear();
  }
  renderMerchantList();
  loadCompetitorAds();
}

function updateStats(stats) {
  document.getElementById('stat-total-ads').textContent = stats.total_ads || 0;
  document.getElementById('stat-merchants').textContent = stats.merchant_count || 0;
  document.getElementById('stat-latest').textContent = stats.latest_scraped || '-';
}

async function loadCompetitorAds() {
  const keyword = document.getElementById('search-input').value.trim();
  const merchantIds = Array.from(selectedMerchants);
  
  try {
    const params = new URLSearchParams();
    if (keyword) params.append('keyword', keyword);
    if (merchantIds.length > 0) params.append('merchants', merchantIds.join(','));
    
    const res = await fetch('/api/competitor/ads?' + params.toString());
    const data = await res.json();
    
    if (!data.ok) {
      toast(data.msg || '加载失败', 'error');
      return;
    }
    
    allAds = data.ads || [];
    renderAdsGrid();
  } catch (e) {
    toast(e.message, 'error');
  }
}

function renderAdsGrid() {
  const container = document.getElementById('ads-grid');
  if (allAds.length === 0) {
    container.innerHTML = '<div class="empty-state-comp" style="grid-column:1/-1;"><div class="empty-state-comp-icon">🔍</div><div>未找到匹配的文案</div></div>';
    return;
  }
  
  try {
    container.innerHTML = allAds.map(ad => {
      // headlines/descriptions 是 JSON 字符串，需要解析
      let headlines = [];
      let descriptions = [];
      try { headlines = typeof ad.headlines === 'string' ? JSON.parse(ad.headlines) : (ad.headlines || []); } catch(e) {}
      try { descriptions = typeof ad.descriptions === 'string' ? JSON.parse(ad.descriptions) : (ad.descriptions || []); } catch(e) {}
      
      const headlineText = (headlines[0] && headlines[0].text) ? headlines[0].text : (ad.headline || ad.title || '无标题');
      const descText = (descriptions[0] && descriptions[0].text) ? descriptions[0].text : (ad.description || '无描述');
      
      const safeMerchant = compHtmlEsc(ad.merchant_name || '未知商户');
      const safeHeadline = compHtmlEsc(headlineText);
      const safeDesc = compHtmlEsc(descText);
      // 所有标题/描述拼接用于复制，存入 data-* 属性
      const allHeadlines = headlines.map(h => h.text || '').filter(Boolean).join(' | ');
      const allDescs = descriptions.map(d => d.text || '').filter(Boolean).join(' | ');
      const safeCopyH = compHtmlEsc(allHeadlines || headlineText);
      const safeCopyD = compHtmlEsc(allDescs || descText);
      
      return `<div class="ad-card">
        <div class="ad-card-header">
          <span class="ad-card-source">🏢 ${safeMerchant}</span>
          <span class="ad-card-date">${compHtmlEsc(ad.scraped_at || '')}</span>
        </div>
        <div class="ad-card-headline">${safeHeadline}</div>
        <div class="ad-card-desc">${safeDesc}</div>
        <div class="ad-card-footer">
          <span class="ad-card-tag">QS: ${ad.quality_score ? Math.round(ad.quality_score) : '-'}</span>
          <span class="ad-card-tag">${compHtmlEsc(ad.asin || '')}</span>
          <button class="copy-btn" data-headline="${safeCopyH}" data-desc="${safeCopyD}" onclick="copyAd(this)">📋 复制</button>
        </div>
      </div>`;
    }).join('');
  } catch(renderErr) {
    console.error('renderAdsGrid 渲染失败:', renderErr);
    container.innerHTML = '<div class="empty-state-comp" style="grid-column:1/-1;color:#ef5350;">渲染失败: ' + compHtmlEsc(renderErr.message) + '</div>';
  }
}

function escapeHtml(text) {
  return compHtmlEsc(text);
}

async function copyAd(btn) {
  // 通过 data-* 属性安全获取内容（避免引号嵌套）
  const headline = btn.dataset.headline || '';
  const desc = btn.dataset.desc || '';
  // data-* 属性里的内容是 HTML 实体，需要解码回原始字符串
  const tmp = document.createElement('textarea');
  tmp.innerHTML = headline;
  const rawH = tmp.value;
  tmp.innerHTML = desc;
  const rawD = tmp.value;
  const text = rawH + String.fromCharCode(10) + rawD;
  try {
    await navigator.clipboard.writeText(text);
    btn.textContent = '✅ 已复制';
    btn.classList.add('copied');
    setTimeout(() => {
      btn.textContent = '📋 复制';
      btn.classList.remove('copied');
    }, 2000);
  } catch (e) {
    toast('复制失败', 'error');
  }
}

function handleSearch() {
  clearTimeout(window.searchTimeout);
  window.searchTimeout = setTimeout(loadCompetitorAds, 300);
}

function clearSearch() {
  document.getElementById('search-input').value = '';
  loadCompetitorAds();
}

window.onload = function() {
  loadMerchants();
  loadCompetitorAds();
};
</script>
</body>
</html>
"""
)


@bp.route("/competitor_ads")
def competitor_ads_page():
    return render_template_string(COMPETITOR_ADS_HTML)


# ═══════════════════════════════════════════════════════════════════════════
# 补充缺失 API（修复前端 JSON 解析错误：Unexpected token '<'）
# ═══════════════════════════════════════════════════════════════════════════


@bp.route("/api/qs/dashboard", methods=["GET"])
def api_qs_dashboard():
    """QS 仪表板数据（转发到 /api/ads/scores，与前端 /qs_dashboard 对接）"""
    try:
        _ensure_quality_score_column()
        conn = get_db()
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT p.asin, p.product_name, p.avg_quality_score,
                   p.ad_count,
                   COUNT(a.id) AS scored_ads,
                   SUM(CASE WHEN a.quality_score >= 80 THEN 1 ELSE 0 END) AS qs_good,
                   SUM(CASE WHEN a.quality_score >= 60 AND a.quality_score < 80 THEN 1 ELSE 0 END) AS qs_ok,
                   SUM(CASE WHEN a.quality_score < 60 THEN 1 ELSE 0 END) AS qs_bad
            FROM ads_plans p
            LEFT JOIN ads_ads a ON a.asin = p.asin
            WHERE p.plan_status = 'completed'
            GROUP BY p.asin, p.product_name, p.avg_quality_score, p.ad_count
            ORDER BY p.avg_quality_score ASC
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        total_ads = sum(r["scored_ads"] or 0 for r in rows)
        qs_good_total = sum(r["qs_good"] or 0 for r in rows)
        qs_bad_total = sum(r["qs_bad"] or 0 for r in rows)
        avg_overall = (
            sum((r["avg_quality_score"] or 0) for r in rows) / len(rows) if rows else 0
        )
        return jsonify(
            {
                "ok": True,
                "avg_qs": round(avg_overall, 1),
                "qs_good_count": qs_good_total,
                "qs_bad_count": qs_bad_total,
                "total_ads": total_ads,
                "total_asins": len(rows),
                "items": rows,
            }
        )
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


@bp.route("/api/competitor/merchants", methods=["GET"])
def api_competitor_merchants():
    """竞品商户列表（用于 /competitor_ads 侧边栏）"""
    try:
        conn = get_db()
        cur = conn.cursor(dictionary=True)
        # ads_plans.merchant_id 是 YP 平台 ID（字符串），与 yp_merchants.id（自增整数）不一致
        # 直接从 ads_plans 聚合有广告方案的商户信息即可
        cur.execute("""
            SELECT p.merchant_id AS id,
                   p.merchant_name AS name,
                   COUNT(DISTINCT a.id) AS ad_count
            FROM ads_plans p
            JOIN ads_ads a ON a.asin = p.asin
            WHERE p.plan_status = 'completed'
            GROUP BY p.merchant_id, p.merchant_name
            ORDER BY ad_count DESC
        """)
        merchants = cur.fetchall()
        cur.close()
        conn.close()
        return jsonify(
            {"ok": True, "merchants": merchants, "stats": {"total": len(merchants)}}
        )
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


@bp.route("/api/competitor/ads", methods=["GET"])
def api_competitor_ads():
    """竞品广告列表（用于 /competitor_ads 主内容区）"""
    try:
        merchant_ids_str = request.args.get("merchants", "")
        search = request.args.get("search", "").strip()
        page = int(request.args.get("page", 1))
        size = int(request.args.get("size", 20))
        offset = (page - 1) * size

        conn = get_db()
        cur = conn.cursor(dictionary=True)

        # merchant_name 直接取 ads_plans，不 JOIN yp_merchants（两表 id 字段不匹配）
        sql = """
            SELECT a.id, a.asin, a.headlines, a.descriptions, a.quality_score,
                   p.product_name, p.merchant_name AS merchant_name
            FROM ads_ads a
            JOIN ads_plans p ON p.asin = a.asin
            WHERE 1=1
        """
        params = []

        if merchant_ids_str:
            # merchant_id 是字符串类型的 YP 平台 ID
            merchant_ids = [x.strip() for x in merchant_ids_str.split(",") if x.strip()]
            if merchant_ids:
                placeholders = ",".join(["%s"] * len(merchant_ids))
                sql += f" AND p.merchant_id IN ({placeholders})"
                params.extend(merchant_ids)

        if search:
            sql += " AND (JSON_SEARCH(a.headlines, 'one', %s) IS NOT NULL OR JSON_SEARCH(a.descriptions, 'one', %s) IS NOT NULL)"
            params.extend([f"%{search}%", f"%{search}%"])

        count_sql = f"SELECT COUNT(*) AS cnt FROM ({sql}) AS _sub"
        cur.execute(count_sql, params)
        total = cur.fetchone()["cnt"]

        sql += " ORDER BY a.quality_score DESC LIMIT %s OFFSET %s"
        params.extend([size, offset])
        cur.execute(sql, params)
        ads = cur.fetchall()
        # headlines/descriptions 是 JSON 字段，已自动解析为 Python 对象
        cur.close()
        conn.close()

        return jsonify(
            {
                "ok": True,
                "ads": ads,
                "total": total,
                "page": page,
                "size": size,
                "pages": max(1, (total + size - 1) // size),
            }
        )
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════
# YP 全量商户同步控制
# ═══════════════════════════════════════════════════════════════════════════


def _get_yp_sync_state():
    """读取同步状态"""
    if YP_SYNC_STATE.exists():
        try:
            return json.loads(YP_SYNC_STATE.read_text(encoding="utf-8"))
        except:
            pass
    return {}


def _is_yp_sync_running():
    """检查同步进程是否在运行"""
    global _yp_sync_proc
    if _yp_sync_proc and _yp_sync_proc.poll() is None:
        return True
    # 也检查系统进程
    try:
        import wmi

        for p in wmi.WMI().Win32_Process():
            if (
                "yp_sync_merchants" in (p.CommandLine or "")
                and p.Name == "python3.12.exe"
            ):
                return True
    except:
        pass
    return False


@bp.route("/yp_sync")
def page_yp_sync():
    return YP_SYNC_HTML


@bp.route("/api/yp_sync/status")
def api_yp_sync_status():
    state = _get_yp_sync_state()
    running = _is_yp_sync_running()
    db_count = 0
    db_unique = 0
    try:
        conn = _db()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM yp_merchants")
        db_count = cur.fetchone()[0]
        cur.execute("SELECT COUNT(DISTINCT merchant_id) FROM yp_merchants")
        db_unique = cur.fetchone()[0]
        cur.close()
        conn.close()
    except:
        pass
    # 读日志最后20行
    log_lines = []
    if YP_SYNC_LOG.exists():
        try:
            log_lines = YP_SYNC_LOG.read_text(
                encoding="utf-8", errors="replace"
            ).splitlines()[-20:]
        except:
            pass
    return jsonify(
        {
            "ok": True,
            "running": running,
            "state": state,
            "db_count": db_count,
            "db_unique": db_unique,
            "log_lines": log_lines,
        }
    )


@bp.route("/api/yp_sync/start", methods=["POST"])
def api_yp_sync_start():
    global _yp_sync_proc
    if _is_yp_sync_running():
        return jsonify({"ok": False, "msg": "同步已在运行中"})
    import subprocess

    _yp_sync_proc = subprocess.Popen(
        [sys.executable, str(YP_SYNC_SCRIPT)],
        cwd=str(BASE_DIR),
        stdout=open(str(YP_SYNC_LOG), "a", encoding="utf-8"),
        stderr=subprocess.STDOUT,
    )
    return jsonify({"ok": True, "msg": "同步已启动"})


@bp.route("/api/yp_sync/stop", methods=["POST"])
def api_yp_sync_stop():
    # 写入停止标记
    import subprocess

    try:
        # kill 掉正在跑的 python3.12 yp_sync_merchants 进程
        result = subprocess.run(
            'taskkill /F /FI "WINDOWTITLE eq yp_sync*" /T 2>nul',
            shell=True,
            capture_output=True,
            text=True,
        )
        result = subprocess.run(
            "wmic process where \"CommandLine like '%%yp_sync_merchants%%' and Name='python3.12.exe'\" call terminate",
            shell=True,
            capture_output=True,
            text=True,
        )
    except:
        pass
    return jsonify({"ok": True, "msg": "已发送停止信号"})


@bp.route("/api/yp_sync/reset", methods=["POST"])
def api_yp_sync_reset():
    if _is_yp_sync_running():
        return jsonify({"ok": False, "msg": "请先停止同步再重置"})
    if YP_SYNC_STATE.exists():
        YP_SYNC_STATE.write_text(
            json.dumps(
                {
                    "page": 0,
                    "total": 0,
                    "total_saved": 0,
                    "started_at": None,
                    "last_run_at": None,
                }
            ),
            encoding="utf-8",
        )
    return jsonify({"ok": True, "msg": "进度已重置"})


@bp.route("/api/yp_sync/collect_mid", methods=["POST"])
def api_yp_sync_collect_mid():
    """按商户ID列表采集商户数据（通过 yp_sync_merchants.py --mid）"""
    import subprocess

    data = request.get_json(silent=True) or {}
    merchant_ids = data.get("merchant_ids", [])
    if not merchant_ids:
        return jsonify({"ok": False, "msg": "请提供 merchant_id"})

    script = str(YP_SYNC_SCRIPT)
    if not os.path.exists(script):
        return jsonify({"ok": False, "msg": f"采集脚本不存在: {script}"})

    log_file = OUTPUT_DIR / "mid_collect_merchant_log.txt"
    try:
        proc = subprocess.Popen(
            [sys.executable, script, "--mid"]
            + [str(m).strip() for m in merchant_ids if str(m).strip()],
            cwd=str(BASE_DIR),
            stdout=open(str(log_file), "a", encoding="utf-8"),
            stderr=subprocess.STDOUT,
        )
        return jsonify(
            {
                "ok": True,
                "msg": f"已启动商户采集：{', '.join(str(m) for m in merchant_ids)}",
                "pid": proc.pid,
            }
        )
    except Exception as e:
        return jsonify({"ok": False, "msg": f"启动失败: {e}"})


YP_SYNC_HTML = (
    "<!DOCTYPE html>\n<html lang='zh-CN'>\n<head>\n<meta charset='utf-8'>\n"
    "<meta name='viewport' content='width=device-width, initial-scale=1'>\n"
    "<title>YP全量同步 · YP Affiliate 管理台</title>\n"
    + _BASE_STYLE_DARK
    + """
<style>
.stat-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(160px,1fr));gap:14px;margin-bottom:24px;}
.stat-card{background:#1a1d24;border:1px solid #2a2d36;border-radius:10px;padding:18px 20px;}
.stat-card .label{font-size:.78rem;color:#888;margin-bottom:6px;text-transform:uppercase;}
.stat-card .value{font-size:1.8rem;font-weight:700;color:#fff;}
.stat-card .value.green{color:#69f0ae;} .stat-card .value.red{color:#f44336;} .stat-card .value.orange{color:#ffb74d;} .stat-card .value.blue{color:#64b5f6;}
.progress-wrap{background:#1a1d24;border:1px solid #2a2d36;border-radius:10px;padding:20px 22px;margin-bottom:24px;}
.progress-label{display:flex;justify-content:space-between;font-size:.83rem;color:#888;margin-bottom:8px;}
.progress-bg{background:#2a2d36;border-radius:8px;height:12px;overflow:hidden;}
.progress-fill{height:100%;border-radius:8px;background:linear-gradient(90deg,#1565c0,#42a5f5);transition:width .5s;}
.btn-row{display:flex;gap:14px;flex-wrap:wrap;align-items:center;margin-bottom:24px;}
.btn-dl{padding:12px 28px;border:none;border-radius:9px;font-size:.95rem;font-weight:600;cursor:pointer;}
.btn-dl:disabled{opacity:.4;cursor:not-allowed;}
.btn-dl-start{background:#2e7d32;color:#fff;} .btn-dl-start:hover:not(:disabled){background:#388e3c;}
.btn-dl-stop{background:#c62828;color:#fff;} .btn-dl-stop:hover:not(:disabled){background:#d32f2f;}
.btn-dl-warning{background:#e65100;color:#fff;} .btn-dl-warning:hover:not(:disabled){background:#f57c00;}
.log-card{background:#1a1d24;border:1px solid #2a2d36;border-radius:10px;padding:18px 20px;}
.log-title{font-size:.82rem;color:#888;text-transform:uppercase;margin-bottom:10px;}
.log-box{background:#0f1117;border-radius:8px;padding:14px;font-family:monospace;font-size:.8rem;color:#ccc;line-height:1.6;min-height:80px;max-height:320px;overflow-y:auto;white-space:pre-wrap;}
.note-card{background:#1a1d24;border:1px solid #2a2d36;border-radius:10px;padding:16px 20px;margin-bottom:24px;font-size:.84rem;color:#aaa;line-height:1.7;}
.note-card b{color:#e0e0e0;} .note-card code{background:#23262f;padding:2px 6px;border-radius:4px;font-family:monospace;font-size:.82rem;color:#64b5f6;}
.status-text{font-size:.88rem;margin-left:8px;}
.toast-container{position:fixed;top:24px;right:24px;z-index:9999;display:flex;flex-direction:column;gap:10px;}
.toast{padding:12px 22px;border-radius:8px;font-size:.88rem;color:#fff;opacity:0;transform:translateX(40px);animation:toastIn .3s forwards;}
.toast-success{background:#2e7d32;}
.toast-error{background:#c62828;}
.toast-info{background:#1565c0;}
@keyframes toastIn{to{opacity:1;transform:translateX(0);}}
</style>
</head>
<body>
"""
    + _SCRAPE_TOPNAV
    + """
<div class="page" style="max-width:900px;">
  <h2 style="font-size:1.5rem;color:#fff;margin-bottom:6px;">🌐 YP 全量商户同步</h2>
  <p style="color:#888;font-size:.88rem;margin-bottom:22px;">从 YP API 分页采集全部商户数据至 MySQL，严格遵守速率限制 (1000条/10分钟)</p>

  <h2 style="font-size:1.3rem;color:#fff;margin-bottom:6px;">🏪 按商户 ID 采集商户数据</h2>
  <p style="color:#888;font-size:.85rem;margin-bottom:18px;">输入 YP 平台的 merchant_id（数字），通过 API 直接采集该商户信息，写入 yp_merchants 表。无需 Chrome。</p>
  <div class="note-card">
    <b>速度极快：</b>直接调用 YP API（advert_id 参数），秒级返回，无需浏览器。<br>
    <b>查找 merchant_id：</b>在「商户管理」页面可查看所有商户及其 ID。<br>
    <b>多次采集安全：</b>使用 INSERT ... ON DUPLICATE KEY UPDATE，重复采集会更新已有记录。
  </div>
  <div style="display:flex;gap:12px;align-items:center;margin-bottom:16px;flex-wrap:wrap;">
    <input type="text" id="midMerchantInput" placeholder="输入 merchant_id，多个用逗号分隔（如 111335,112887）"
      style="flex:1;min-width:280px;padding:10px 16px;background:#0f1117;border:1px solid #2a2d36;border-radius:8px;color:#fff;font-size:.9rem;outline:none;">
    <button class="btn-dl btn-dl-start" id="btnMidMerchantCollect" onclick="collectMerchantByMid()">▶ 开始采集</button>
  </div>
  <div class="log-card">
    <div class="log-title">采集日志</div>
    <div class="log-box" id="midMerchantResultBox" style="min-height:50px;">等待操作...</div>
  </div>
</div>
<div class="page" style="max-width:900px;margin-top:30px;">
  <div class="stat-grid">
    <div class="stat-card"><div class="label">YP 平台商户总数</div><div class="value blue" id="s-total">-</div></div>
    <div class="stat-card"><div class="label">本次已同步</div><div class="value green" id="s-saved">-</div></div>
    <div class="stat-card"><div class="label">数据库记录数</div><div class="value" id="s-db">-</div></div>
    <div class="stat-card"><div class="label">唯一商户数</div><div class="value blue" id="s-unique">-</div></div>
    <div class="stat-card"><div class="label">当前页/总页数</div><div class="value orange" id="s-page">-</div></div>
  </div>

  <div class="progress-wrap">
    <div class="progress-label"><span id="s-percent">同步进度: -</span><span id="s-time">-</span></div>
    <div class="progress-bg"><div class="progress-fill" id="s-bar" style="width:0%"></div></div>
  </div>

  <div class="btn-row">
    <button class="btn-dl btn-dl-start" id="btn-start" onclick="startSync()">▶ 开始同步</button>
    <button class="btn-dl btn-dl-stop" id="btn-stop" onclick="stopSync()" disabled>⏹ 暂停</button>
    <button class="btn-dl btn-dl-warning" id="btn-reset" onclick="resetSync()">🔄 重置进度</button>
    <span class="status-text" id="s-running" style="color:#f44336;">⏹ 已停止</span>
  </div>

  <div class="log-card">
    <div class="log-title">📋 实时日志</div>
    <div class="log-box" id="log-box">等待日志...</div>
  </div>

  <div class="note-card">
    <b>速率限制：</b>每页采集 <code>1000</code> 条，批次间隔 <code>10 分钟</code>（遵守 YP 速率限制）<br>
    <b>断点续传：</b>暂停/停止后再次启动会从上次位置继续<br>
    <b>写入方式：</b>全量 INSERT，保留所有原始记录（不去重）<br>
    <b>自动重置：</b>同步完成后状态自动重置，下次启动为全量检查
  </div>
</div>
<script>
var refreshTimer = null;
var isRunning = false;

function updateUI(data) {
  var s = data.state || {};
  isRunning = data.running;

  document.getElementById('s-total').textContent = (s.total || 0).toLocaleString();
  document.getElementById('s-saved').textContent = (s.total_saved || 0).toLocaleString();
  document.getElementById('s-db').textContent = (data.db_count || 0).toLocaleString();
  document.getElementById('s-unique').textContent = (data.db_unique || 0).toLocaleString();

  var page = s.page || 0;
  var total = s.total || 0;
  var total_pages = Math.ceil(total / 1000);
  document.getElementById('s-page').textContent = page + ' / ' + total_pages;

  var pct = total > 0 ? (page / total_pages * 100).toFixed(1) : 0;
  document.getElementById('s-percent').textContent = '同步进度: ' + pct + '%';
  document.getElementById('s-bar').style.width = pct + '%';

  // 估算剩余时间
  if (isRunning && page > 0) {
    var remaining = (total_pages - page) * 10 / 60;
    document.getElementById('s-time').textContent = '预计剩余: ' + remaining.toFixed(1) + ' 小时';
  } else if (page >= total_pages && total_pages > 0) {
    document.getElementById('s-time').textContent = '✅ 已完成';
  } else {
    document.getElementById('s-time').textContent = '-';
  }

  // 按钮状态
  document.getElementById('btn-start').disabled = isRunning;
  document.getElementById('btn-stop').disabled = !isRunning;
  document.getElementById('btn-reset').disabled = isRunning;

  var statusEl = document.getElementById('s-running');
  if (isRunning) {
    statusEl.innerHTML = '<span class=\"badge badge-running\">同步中...</span>';
    statusEl.style.color = '#4caf50';
  } else {
    statusEl.textContent = '⏹ 已停止';
    statusEl.style.color = '#f44336';
  }

  // 日志
  if (data.log_lines && data.log_lines.length) {
    var box = document.getElementById('log-box');
    box.innerHTML = data.log_lines.map(function(l) {
      return '<div>' + l.replace(/</g,'&lt;') + '</div>';
    }).join('');
    box.scrollTop = box.scrollHeight;
  }
}

function refresh() {
  fetch('/api/yp_sync/status').then(function(r){return r.json()}).then(function(d){
    updateUI(d);
  }).catch(function(){});
}

function startSync() {
  fetch('/api/yp_sync/start', {method:'POST'}).then(function(r){return r.json()}).then(function(d){
    toast(d.msg, d.ok?'success':'error');
    setTimeout(refresh, 1000);
  });
}

function stopSync() {
  if (!confirm('确定要暂停同步吗？进度已保存，稍后可继续。')) return;
  fetch('/api/yp_sync/stop', {method:'POST'}).then(function(r){return r.json()}).then(function(d){
    toast(d.msg, d.ok?'success':'error');
    setTimeout(refresh, 2000);
  });
}

function resetSync() {
  if (!confirm('确定要重置进度吗？下次将从头开始。')) return;
  fetch('/api/yp_sync/reset', {method:'POST'}).then(function(r){return r.json()}).then(function(d){
    toast(d.msg, d.ok?'success':'error');
    setTimeout(refresh, 500);
  });
}

function toast(msg, type) {
  type = type || 'info';
  var c = document.getElementById('toast-container');
  var t = document.createElement('div');
  t.className = 'toast toast-' + type;
  t.textContent = msg;
  c.appendChild(t);
  setTimeout(function() { t.style.opacity='0'; t.style.transform='translateX(40px)'; setTimeout(function(){t.remove();},300); }, 3000);
}

// 启动定时刷新
refresh();
refreshTimer = setInterval(refresh, 5000);

function collectMerchantByMid() {
  var input = document.getElementById('midMerchantInput').value.trim();
  if (!input) { toast('请输入 merchant_id', 'error'); return; }
  var mids = input.split(/[,，\s]+/).filter(function(s) { return s.length > 0; });
  var resultBox = document.getElementById('midMerchantResultBox');
  var btn = document.getElementById('btnMidMerchantCollect');
  btn.disabled = true;
  btn.textContent = '启动中...';
  resultBox.textContent = '正在采集商户数据...';
  fetch('/api/yp_sync/collect_mid', {
    method: 'POST',
    headers: { 'Content-Type': 'application/json' },
    body: JSON.stringify({ merchant_ids: mids })
  }).then(function(r) { return r.json(); }).then(function(d) {
    btn.disabled = false;
    btn.textContent = '\u25b6 开始采集';
    if (d.ok) {
      resultBox.textContent = '[' + new Date().toLocaleTimeString() + '] ' + d.msg + '\\n\\n提示：采集通过 YP API 直接完成，通常几秒内写入数据库。刷新页面可查看最新数据。';
      toast('商户采集已启动', 'success');
      setTimeout(refresh, 3000);
    } else {
      resultBox.textContent = 'ERROR: ' + (d.msg || '未知错误');
      toast(d.msg || '启动失败', 'error');
    }
  }).catch(function(e) {
    btn.disabled = false;
    btn.textContent = '\u25b6 开始采集';
    resultBox.textContent = '请求失败: ' + e.message;
    toast('请求失败', 'error');
  });
}
</script>
<div class="toast-container" id="toast-container"></div>
</body>
</html>
"""
)
